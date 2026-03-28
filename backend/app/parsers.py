import os
import re
from typing import Any

import pdfplumber
from openpyxl import load_workbook


def _clean(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _parse_bool(value: Any) -> bool | None:
    if value is None:
        return None
    raw = _clean(value)
    if raw == "":
        return False

    lowered = raw.lower()
    true_values = {
        "o",
        "0",
        "y",
        "yes",
        "true",
        "t",
        "v",
        "ok",
        "checked",
        "완료",
        "예",
        "유",
        "있음",
        "당원",
        "✓",
        "✔",
        "☑",
        "○",
        "●",
    }
    false_values = {
        "x",
        "n",
        "no",
        "false",
        "f",
        "미완료",
        "아니오",
        "무",
        "없음",
        "비당원",
        "미통화",
        "미체크",
        "✗",
        "×",
    }

    if lowered in true_values or raw in true_values:
        return True
    if lowered in false_values or raw in false_values:
        return False

    if lowered.isdigit():
        return int(lowered) > 0

    return None


def _guess_owner_name_from_filename(filename: str) -> str:
    stem = os.path.splitext(os.path.basename(filename))[0]
    match = re.search(r"-\s*(.+)$", stem)
    if match:
        return match.group(1).strip()
    return stem


def _extract_owner_name(text: str) -> str | None:
    patterns = [
        r"제공자\s*명\s*[:：]?\s*([가-힣A-Za-z0-9]{2,30})",
        r"제공자\s*[:：]?\s*([가-힣A-Za-z0-9]{2,30})",
    ]
    for pattern in patterns:
        match = re.search(pattern, text)
        if match:
            return match.group(1).strip()
    return None


def _extract_owner_phone(text: str) -> str | None:
    patterns = [
        r"제공자\s*전화번호\s*[:：]?\s*([0-9\-]{8,20})",
        r"제공자\s*연락처\s*[:：]?\s*([0-9\-]{8,20})",
        r"연락처\s*[:：]?\s*([0-9\-]{8,20})",
    ]
    for pattern in patterns:
        match = re.search(pattern, text)
        if match:
            return match.group(1)
    return None


def _header_index_map(headers: list[str]) -> dict[str, int]:
    index: dict[str, int] = {}
    for idx, name in enumerate(headers):
        col = name.replace(" ", "")
        if "이름" in col and "name" not in index:
            index["name"] = idx
        if ("전화번호" in col or "연락처" in col) and "phone" not in index:
            index["phone"] = idx
        if ("친밀" in col or "친밀도" in col) and "intimacy" not in index:
            index["intimacy"] = idx
        if ("전화여부" in col or "통화" in col) and "called" not in index:
            index["called"] = idx
        if "당원" in col and "party" not in index:
            index["party"] = idx

    if "phone" not in index:
        for idx, name in enumerate(headers):
            col = name.replace(" ", "")
            if "전화" in col and "여부" not in col:
                index["phone"] = idx
                break

    return index


def parse_excel_file(path: str, original_filename: str) -> dict[str, Any]:
    wb = load_workbook(path, data_only=True)
    ws = wb.active

    owner_name = _clean(ws["B1"].value) or _guess_owner_name_from_filename(original_filename)
    owner_phone = None

    header_row = None
    header_map: dict[str, int] = {}

    for row_number in range(1, min(ws.max_row, 30) + 1):
        headers = [_clean(ws.cell(row=row_number, column=col).value) for col in range(1, 30)]
        candidate = _header_index_map(headers)
        if "name" in candidate and "phone" in candidate:
            header_row = row_number
            header_map = candidate
            break

    if header_row is None:
        # Fallback to a common layout when header detection fails.
        header_row = 2
        header_map = {"name": 0, "phone": 1, "intimacy": 2, "called": 3, "party": 4}

    records: list[dict[str, Any]] = []
    owner_phone_candidates: list[str] = []
    for row_number in range(header_row + 1, ws.max_row + 1):
        row_values = [_clean(ws.cell(row=row_number, column=col).value) for col in range(1, 30)]
        name = row_values[header_map["name"]] if header_map.get("name") is not None else ""
        phone = row_values[header_map["phone"]] if header_map.get("phone") is not None else ""

        if not name and not phone:
            continue

        # Excel rule: owner phone is discovered from rows where name == owner_name.
        if owner_name and name == owner_name and phone:
            owner_phone_candidates.append(phone)

        record = {
            "name": name,
            "phone": phone,
            "intimacy_checked": _parse_bool(
                row_values[header_map["intimacy"]] if header_map.get("intimacy") is not None else None
            ),
            "called": _parse_bool(
                row_values[header_map["called"]] if header_map.get("called") is not None else None
            ),
            "party_member": _parse_bool(
                row_values[header_map["party"]] if header_map.get("party") is not None else None
            ),
        }
        records.append(record)

    unique_owner_phones = {
        normalize_mobile_phone(phone) for phone in owner_phone_candidates if normalize_mobile_phone(phone)
    }
    if len(unique_owner_phones) == 1:
        owner_phone = format_phone(next(iter(unique_owner_phones)))
    else:
        # If owner name appears multiple times with different phones, keep owner phone blank.
        owner_phone = None

    return {
        "owner_name": owner_name,
        "owner_phone": owner_phone,
        "records": records,
    }


def parse_pdf_file(path: str, original_filename: str) -> dict[str, Any]:
    owner_name = _guess_owner_name_from_filename(original_filename)
    owner_phone = None
    records: list[dict[str, Any]] = []

    with pdfplumber.open(path) as pdf:
        first_page_text = pdf.pages[0].extract_text() or "" if pdf.pages else ""
        owner_name = _extract_owner_name(first_page_text) or owner_name
        owner_phone = _extract_owner_phone(first_page_text)

        for page in pdf.pages:
            for table in page.extract_tables() or []:
                header_map: dict[str, int] | None = None
                for row in table:
                    cleaned = [_clean(cell) for cell in row]
                    if not any(cleaned):
                        continue

                    candidate_header = _header_index_map(cleaned)
                    if "name" in candidate_header and "phone" in candidate_header:
                        header_map = candidate_header
                        continue

                    if header_map is None:
                        continue

                    name = cleaned[header_map["name"]] if header_map.get("name") is not None else ""
                    phone = cleaned[header_map["phone"]] if header_map.get("phone") is not None else ""
                    if not name and not phone:
                        continue

                    records.append(
                        {
                            "name": name,
                            "phone": phone,
                            "intimacy_checked": _parse_bool(
                                cleaned[header_map["intimacy"]]
                                if header_map.get("intimacy") is not None
                                else None
                            ),
                            "called": _parse_bool(
                                cleaned[header_map["called"]]
                                if header_map.get("called") is not None
                                else None
                            ),
                            "party_member": _parse_bool(
                                cleaned[header_map["party"]]
                                if header_map.get("party") is not None
                                else None
                            ),
                        }
                    )

    return {
        "owner_name": owner_name,
        "owner_phone": owner_phone,
        "records": records,
    }


def normalize_phone(value: Any) -> str:
    raw = _clean(value)
    digits = "".join(ch for ch in raw if ch.isdigit())
    if not digits:
        return ""
    return digits


def normalize_mobile_phone(value: Any) -> str:
    digits = normalize_phone(value)
    if len(digits) == 10 and digits.startswith("10"):
        digits = f"0{digits}"
    if len(digits) == 11 and digits.startswith("010"):
        return digits
    return ""


def format_phone(value: str) -> str:
    digits = normalize_phone(value)
    if len(digits) == 11:
        return f"{digits[:3]}-{digits[3:7]}-{digits[7:]}"
    if len(digits) == 10:
        return f"{digits[:3]}-{digits[3:6]}-{digits[6:]}"
    return value


def parse_compare_excel_file(path: str) -> list[dict[str, Any]]:
    wb = load_workbook(path, data_only=True)
    ws = wb.active

    header_aliases = {
        "full_name": ["성명", "이름"],
        "birth_date": ["생년월일", "생년월", "생일"],
        "phone": ["연락처", "전화번호", "휴대폰", "핸드폰", "전화"],
        "province": ["도"],
        "city_county": ["시(군)", "시군", "지역", "시", "군"],
        "district": ["구"],
        "dong": ["동"],
        "address_detail": ["주소 (상세)", "주소(상세)", "상세주소", "주소"],
    }

    header_row = 1
    for row_number in range(1, min(ws.max_row, 20) + 1):
        row_values = [_clean(ws.cell(row=row_number, column=col).value) for col in range(1, 40)]
        normalized = [value.replace(" ", "") for value in row_values]
        if "성명" in normalized and ("연락처" in normalized or "전화번호" in normalized):
            header_row = row_number
            break

    header_values = [_clean(ws.cell(row=header_row, column=col).value) for col in range(1, 40)]
    normalized_headers = [value.replace(" ", "") for value in header_values]
    index_by_key: dict[str, int] = {}
    for key, aliases in header_aliases.items():
        for alias in aliases:
            alias_norm = alias.replace(" ", "")
            if alias_norm in normalized_headers:
                index_by_key[key] = normalized_headers.index(alias_norm)
                break

    rows: list[dict[str, Any]] = []
    for row_number in range(header_row + 1, ws.max_row + 1):
        values = [_clean(ws.cell(row=row_number, column=col).value) for col in range(1, 40)]
        if not any(values):
            continue

        record = {
            "full_name": values[index_by_key.get("full_name", -1)] if "full_name" in index_by_key else "",
            "birth_date": values[index_by_key.get("birth_date", -1)] if "birth_date" in index_by_key else "",
            "phone": values[index_by_key.get("phone", -1)] if "phone" in index_by_key else "",
            "province": values[index_by_key.get("province", -1)] if "province" in index_by_key else "",
            "city_county": values[index_by_key.get("city_county", -1)] if "city_county" in index_by_key else "",
            "district": values[index_by_key.get("district", -1)] if "district" in index_by_key else "",
            "dong": values[index_by_key.get("dong", -1)] if "dong" in index_by_key else "",
            "address_detail": values[index_by_key.get("address_detail", -1)] if "address_detail" in index_by_key else "",
            "_row_number": row_number,
        }
        rows.append(record)

    return rows


def parse_supporter_excel_file(path: str) -> list[dict[str, Any]]:
    wb = load_workbook(path, data_only=True)
    ws = wb.active

    header_row = None
    for row_number in range(1, min(ws.max_row, 20) + 1):
        first = _clean(ws.cell(row=row_number, column=1).value).replace(" ", "")
        second = _clean(ws.cell(row=row_number, column=2).value).replace(" ", "")
        if first in {"이름", "성명", "name"} and second in {"전화번호", "연락처", "휴대폰", "핸드폰", "phone"}:
            header_row = row_number
            break

    start_row = (header_row + 1) if header_row else 1
    rows: list[dict[str, Any]] = []
    for row_number in range(start_row, ws.max_row + 1):
        name = _clean(ws.cell(row=row_number, column=1).value)
        phone = _clean(ws.cell(row=row_number, column=2).value)
        if not name and not phone:
            continue
        rows.append(
            {
                "name": name,
                "phone": phone,
                "_row_number": row_number,
            }
        )

    return rows


def _normalize_jeonju_phone(value: Any) -> str:
    digits = normalize_mobile_phone(value)
    return format_phone(digits) if digits else ""


def _split_korean_address(address: str) -> tuple[str, str, str, str, str]:
    raw = _clean(address)
    if not raw:
        return "", "", "", "", ""

    tokens = raw.split()
    province = ""
    city_county = ""
    district = ""
    dong = ""
    rest_tokens: list[str] = []

    idx = 0
    province_aliases = {
        "서울",
        "부산",
        "대구",
        "인천",
        "광주",
        "대전",
        "울산",
        "세종",
        "경기",
        "강원",
        "충북",
        "충남",
        "전북",
        "전남",
        "경북",
        "경남",
        "제주",
    }

    if idx < len(tokens) and (
        tokens[idx].endswith("도")
        or tokens[idx].endswith("특별시")
        or tokens[idx].endswith("광역시")
        or tokens[idx].endswith("특별자치시")
        or tokens[idx].endswith("특별자치도")
        or tokens[idx] in province_aliases
    ):
        province = tokens[idx]
        idx += 1

    if idx < len(tokens) and (tokens[idx].endswith("시") or tokens[idx].endswith("군")):
        city_county = tokens[idx]
        idx += 1

    if idx < len(tokens) and tokens[idx].endswith("구"):
        district = tokens[idx]
        idx += 1

    if idx < len(tokens) and (
        tokens[idx].endswith("동")
        or tokens[idx].endswith("읍")
        or tokens[idx].endswith("면")
        or tokens[idx].endswith("리")
        or tokens[idx].endswith("가")
    ):
        dong = tokens[idx]
        idx += 1

    if idx < len(tokens):
        rest_tokens = tokens[idx:]

    return province, city_county, district, dong, " ".join(rest_tokens)


def parse_jeonju_upload_excel_file(path: str) -> list[dict[str, Any]]:
    wb = load_workbook(path, data_only=True)

    rows: list[dict[str, Any]] = []
    for ws in wb.worksheets:
        header_row = None
        idx_name = None
        idx_phone = None
        idx_address = None

        for row_number in range(1, min(ws.max_row, 30) + 1):
            values = [_clean(ws.cell(row=row_number, column=col).value) for col in range(1, 40)]
            normalized = [value.replace(" ", "") for value in values]

            for i, value in enumerate(normalized):
                if idx_name is None and value in {"당원명", "성명", "이름"}:
                    idx_name = i
                if idx_phone is None and value in {"휴대전화번호", "휴대폰", "연락처", "전화번호"}:
                    idx_phone = i
                if idx_address is None and value in {"지번주소", "주소", "주소(상세)", "상세주소"}:
                    idx_address = i

            if idx_name is not None and idx_phone is not None and idx_address is not None:
                header_row = row_number
                break

        if header_row is None or idx_name is None or idx_phone is None or idx_address is None:
            continue

        for row_number in range(header_row + 1, ws.max_row + 1):
            values = [_clean(ws.cell(row=row_number, column=col).value) for col in range(1, 40)]
            if not any(values):
                continue

            full_name = values[idx_name] if idx_name < len(values) else ""
            phone = _normalize_jeonju_phone(values[idx_phone] if idx_phone < len(values) else "")
            address = values[idx_address] if idx_address < len(values) else ""
            if not full_name and not phone and not address:
                continue

            province, city_county, district, dong, address_detail = _split_korean_address(address)
            rows.append(
                {
                    "full_name": full_name,
                    "birth_date": "",
                    "phone": phone,
                    "province": province,
                    "city_county": city_county,
                    "district": district,
                    "dong": dong,
                    "address_detail": address_detail,
                }
            )

    return rows
