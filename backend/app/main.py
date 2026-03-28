import os
import tempfile
import threading
import time
import uuid
from csv import DictWriter
from datetime import datetime
from io import BytesIO
from pathlib import Path
from io import StringIO
from zoneinfo import ZoneInfo

from fastapi import Depends, FastAPI, File, Form, Header, HTTPException, Query, UploadFile, status
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from sqlalchemy import text
from sqlalchemy.orm import Session

from .config import settings
from .database import SessionLocal
from .parsers import (
    format_phone,
    normalize_mobile_phone,
    normalize_phone,
    parse_compare_excel_file,
    parse_excel_file,
    parse_jeonju_upload_excel_file,
    parse_supporter_excel_file,
    parse_pdf_file,
)
from .schemas import (
    ContactCategoryRead,
    ContactListItemRead,
    ContactListRead,
    ContactOwnerRead,
    CompareRecordListRead,
    CompareRecordRead,
    CompareUploadSummary,
    DailyStatsRead,
    GroupCreate,
    GroupOwnerRead,
    GroupRead,
    LoginRequest,
    LoginResponse,
    ElectionContactsRead,
    ElectionDistrictRead,
    ElectionDongAddRequest,
    OwnerDetailRead,
    OwnerRecordRead,
    SupporterListItemRead,
    SupporterListRead,
    SupporterStatsSummaryRead,
    SupporterUploadAsyncQueued,
    SupporterUploadSyncSummary,
    StatsSummaryRead,
    TreeGroupNodeRead,
    TreeOwnerNodeRead,
    TodayManagerRead,
    UnifiedContactListItemRead,
    UnifiedContactListRead,
    UploadFileResult,
    UploadSummary,
)

app = FastAPI(title="doc-diff API", version="0.2.0")

ADMIN_USERNAME = "admin"
ADMIN_PASSWORD = "supporter1@"
ACCESS_TOKEN = "docdiff-admin-token"
KST = ZoneInfo("Asia/Seoul")
_stats_thread_started = False
_stats_thread_lock = threading.Lock()
_upload_jobs: dict[str, dict] = {}
_upload_jobs_lock = threading.Lock()
_supporter_upload_jobs: dict[str, dict] = {}
_supporter_upload_jobs_lock = threading.Lock()
SUPPORTER_MATCHED_SCOPE = "matched"
SUPPORTER_TOTAL_SCOPE = "total"
SUPPORTER_ADDRESS_MATCH_CONDITION = """
    COALESCE(TRIM(c.city_county), '') <> ''
    OR COALESCE(TRIM(c.district), '') <> ''
    OR COALESCE(TRIM(c.dong), '') <> ''
    OR COALESCE(TRIM(c.address_detail), '') <> ''
"""
UNIFIED_SCOPE_TOTAL = "total"
UNIFIED_SCOPE_MATCHED = "matched"

DEFAULT_ELECTION_DISTRICT_DONGS: dict[str, list[str]] = {
    "가선거구": ["중앙동", "풍남동", "노송동", "인후동3가"],
    "나선거구": ["완산동", "중화산동1가", "중화산동2가"],
    "다선거구": ["동서학동", "서서학동", "평화동1가", "평화동2가"],
    "라선거구": ["서신동"],
    "마선거구": ["삼천동1가", "삼천동2가", "삼천동3가", "효자동1가"],
    "바선거구": ["효자동2가", "효자동3가", "효자동4가"],
    "사선거구": ["효자동5가"],
    "아선거구": ["진북동", "인후동1가", "인후동2가", "금암동"],
    "자선거구": ["덕진동", "팔복동", "송천동2가"],
    "차선거구": ["우아동1가", "우아동2가", "호성동"],
    "카선거구": ["송천동1가"],
    "타선거구": ["조촌동", "여의동", "혁신동"],
}

origins = [origin.strip() for origin in settings.cors_origins.split(",") if origin.strip()]
app.add_middleware(
    CORSMiddleware,
    allow_origins=origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()


def require_auth(authorization: str | None = Header(default=None)):
    expected = f"Bearer {ACCESS_TOKEN}"
    if authorization != expected:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Unauthorized",
        )


def ensure_tables(db: Session) -> None:
    db.execute(
        text(
            """
            CREATE TABLE IF NOT EXISTS data_groups (
              id SERIAL PRIMARY KEY,
              name TEXT NOT NULL UNIQUE,
              created_at TIMESTAMPTZ NOT NULL DEFAULT NOW()
            );

            CREATE TABLE IF NOT EXISTS data_owners (
              id SERIAL PRIMARY KEY,
              group_id INTEGER NOT NULL REFERENCES data_groups(id) ON DELETE CASCADE,
              owner_name TEXT NOT NULL,
              owner_phone TEXT,
              owner_phone_normalized TEXT,
              source_type TEXT NOT NULL,
              file_name TEXT NOT NULL,
              uploaded_at TIMESTAMPTZ NOT NULL DEFAULT NOW()
            );

            CREATE TABLE IF NOT EXISTS data_records (
              id SERIAL PRIMARY KEY,
              owner_id INTEGER NOT NULL REFERENCES data_owners(id) ON DELETE CASCADE,
              person_name TEXT,
              phone TEXT,
              phone_normalized TEXT,
              intimacy_checked BOOLEAN,
              called BOOLEAN,
              party_member BOOLEAN,
              created_at TIMESTAMPTZ NOT NULL DEFAULT NOW()
            );

            CREATE TABLE IF NOT EXISTS compare_records (
              id SERIAL PRIMARY KEY,
              full_name TEXT,
              birth_date TEXT,
              phone TEXT NOT NULL,
              phone_normalized TEXT NOT NULL UNIQUE,
              province TEXT,
              city_county TEXT,
              district TEXT,
              dong TEXT,
              address_detail TEXT,
              created_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
              updated_at TIMESTAMPTZ NOT NULL DEFAULT NOW()
            );

            CREATE TABLE IF NOT EXISTS compare_upload_errors (
              id SERIAL PRIMARY KEY,
              batch_id TEXT NOT NULL,
              file_name TEXT NOT NULL,
              row_number INTEGER,
              full_name TEXT,
              phone_raw TEXT,
              reason TEXT NOT NULL,
              created_at TIMESTAMPTZ NOT NULL DEFAULT NOW()
            );

            CREATE TABLE IF NOT EXISTS supporter_records (
              id SERIAL PRIMARY KEY,
              supporter_name TEXT,
              phone TEXT NOT NULL,
              phone_normalized TEXT NOT NULL UNIQUE,
              created_at TIMESTAMPTZ NOT NULL DEFAULT NOW()
            );

            CREATE TABLE IF NOT EXISTS stats_view (
              id INTEGER PRIMARY KEY,
              total_managers INTEGER NOT NULL DEFAULT 0,
              today_representatives INTEGER NOT NULL DEFAULT 0,
              today_added_managers INTEGER NOT NULL DEFAULT 0,
              favorite_contacts INTEGER NOT NULL DEFAULT 0,
              total_contacts INTEGER NOT NULL DEFAULT 0,
              refreshed_at TIMESTAMPTZ NOT NULL DEFAULT NOW()
            );

            CREATE TABLE IF NOT EXISTS compare_records_view_meta (
              id INTEGER PRIMARY KEY,
              total_count INTEGER NOT NULL DEFAULT 0,
              dong_count INTEGER NOT NULL DEFAULT 0,
              latest_updated_at TIMESTAMPTZ,
              refreshed_at TIMESTAMPTZ NOT NULL DEFAULT NOW()
            );
            """
        )
    )
    db.execute(
        text(
            """
            CREATE TABLE IF NOT EXISTS contacts_view (
              phone_normalized TEXT PRIMARY KEY,
              phone TEXT,
              person_name TEXT,
              created_at TIMESTAMPTZ NOT NULL,
              has_favorite BOOLEAN NOT NULL DEFAULT FALSE,
              city_county TEXT,
              dong TEXT,
              address_detail TEXT,
              owner_primary_name TEXT,
              owner_count INTEGER NOT NULL DEFAULT 0,
              updated_at TIMESTAMPTZ NOT NULL DEFAULT NOW()
            );

            CREATE TABLE IF NOT EXISTS compare_records_view (
              id INTEGER PRIMARY KEY,
              full_name TEXT,
              birth_date TEXT,
              phone TEXT NOT NULL,
              phone_normalized TEXT NOT NULL UNIQUE,
              province TEXT,
              city_county TEXT,
              district TEXT,
              dong TEXT,
              address_detail TEXT,
              created_at TIMESTAMPTZ NOT NULL,
              updated_at TIMESTAMPTZ NOT NULL
            );

            CREATE TABLE IF NOT EXISTS election_district_dongs (
              id SERIAL PRIMARY KEY,
              district_name TEXT NOT NULL,
              dong TEXT NOT NULL,
              created_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
              UNIQUE(district_name, dong)
            );

            CREATE TABLE IF NOT EXISTS jeonju_records (
              id SERIAL PRIMARY KEY,
              category TEXT NOT NULL,
              jeonju_name TEXT,
              phone TEXT NOT NULL,
              phone_normalized TEXT NOT NULL,
              created_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
              UNIQUE(category, phone_normalized)
            );
            """
        )
    )
    db.execute(text("CREATE INDEX IF NOT EXISTS idx_jeonju_records_category ON jeonju_records(category)"))
    db.execute(text("ALTER TABLE data_owners ADD COLUMN IF NOT EXISTS owner_phone_normalized TEXT"))
    db.execute(text("ALTER TABLE data_records ADD COLUMN IF NOT EXISTS phone_normalized TEXT"))
    db.execute(
        text(
            """
            UPDATE data_owners
            SET owner_phone_normalized = regexp_replace(COALESCE(owner_phone, ''), '[^0-9]', '', 'g')
            WHERE owner_phone_normalized IS NULL
            """
        )
    )
    db.execute(
        text(
            """
            UPDATE data_records
            SET phone_normalized = regexp_replace(COALESCE(phone, ''), '[^0-9]', '', 'g')
            WHERE phone_normalized IS NULL
            """
        )
    )
    db.execute(text("CREATE INDEX IF NOT EXISTS idx_data_owners_phone_norm ON data_owners(owner_phone_normalized)"))
    db.execute(text("CREATE INDEX IF NOT EXISTS idx_data_records_phone_norm ON data_records(phone_normalized)"))
    db.execute(text("CREATE INDEX IF NOT EXISTS idx_contacts_view_created_at ON contacts_view(created_at DESC)"))
    db.execute(text("CREATE INDEX IF NOT EXISTS idx_contacts_view_city ON contacts_view(city_county)"))
    db.execute(text("CREATE INDEX IF NOT EXISTS idx_contacts_view_dong ON contacts_view(dong)"))
    db.execute(text("CREATE INDEX IF NOT EXISTS idx_contacts_view_favorite ON contacts_view(has_favorite)"))
    db.execute(text("CREATE INDEX IF NOT EXISTS idx_election_district_name ON election_district_dongs(district_name)"))
    db.execute(text("CREATE INDEX IF NOT EXISTS idx_election_dong ON election_district_dongs(dong)"))
    db.execute(text("CREATE INDEX IF NOT EXISTS idx_compare_records_view_updated_at ON compare_records_view(updated_at DESC, id DESC)"))
    db.execute(text("CREATE INDEX IF NOT EXISTS idx_compare_records_view_phone_norm ON compare_records_view(phone_normalized)"))
    db.execute(text("CREATE INDEX IF NOT EXISTS idx_supporter_records_created_at ON supporter_records(created_at DESC)"))
    db.execute(text("ALTER TABLE compare_records_view_meta ADD COLUMN IF NOT EXISTS dong_count INTEGER NOT NULL DEFAULT 0"))
    for district_name, dongs in DEFAULT_ELECTION_DISTRICT_DONGS.items():
        for dong in dongs:
            db.execute(
                text(
                    """
                    INSERT INTO election_district_dongs(district_name, dong)
                    VALUES (:district_name, :dong)
                    ON CONFLICT (district_name, dong) DO NOTHING
                    """
                ),
                {"district_name": district_name, "dong": dong},
            )
    db.commit()


def recompute_stats_view(db: Session) -> None:
    row = db.execute(
        text(
            """
            WITH manager_stats AS (
              SELECT
                COUNT(*)::int AS total_managers,
                COUNT(*) FILTER (
                  WHERE (uploaded_at AT TIME ZONE 'Asia/Seoul')::date = (NOW() AT TIME ZONE 'Asia/Seoul')::date
                )::int AS today_added_managers
              FROM data_owners
            ),
            rep_stats AS (
              SELECT
                COUNT(*) FILTER (
                  WHERE (created_at AT TIME ZONE 'Asia/Seoul')::date = (NOW() AT TIME ZONE 'Asia/Seoul')::date
                )::int AS today_representatives
              FROM data_groups
            ),
            contact_stats AS (
              SELECT
                COUNT(DISTINCT phone_normalized) FILTER (
                  WHERE COALESCE(phone_normalized, '') <> ''
                )::int AS total_contacts,
                COUNT(DISTINCT phone_normalized) FILTER (
                  WHERE COALESCE(phone_normalized, '') <> '' AND intimacy_checked = TRUE
                )::int AS favorite_contacts
              FROM data_records
            )
            SELECT
              manager_stats.total_managers,
              rep_stats.today_representatives,
              manager_stats.today_added_managers,
              contact_stats.favorite_contacts,
              contact_stats.total_contacts
            FROM manager_stats, rep_stats, contact_stats
            """
        )
    ).mappings().first()

    db.execute(
        text(
            """
            INSERT INTO stats_view(
              id, total_managers, today_representatives, today_added_managers, favorite_contacts, total_contacts, refreshed_at
            )
            VALUES(
              1, :total_managers, :today_representatives, :today_added_managers, :favorite_contacts, :total_contacts, NOW()
            )
            ON CONFLICT (id)
            DO UPDATE SET
              total_managers = EXCLUDED.total_managers,
              today_representatives = EXCLUDED.today_representatives,
              today_added_managers = EXCLUDED.today_added_managers,
              favorite_contacts = EXCLUDED.favorite_contacts,
              total_contacts = EXCLUDED.total_contacts,
              refreshed_at = NOW()
            """
        ),
        {
            "total_managers": row["total_managers"],
            "today_representatives": row["today_representatives"],
            "today_added_managers": row["today_added_managers"],
            "favorite_contacts": row["favorite_contacts"],
            "total_contacts": row["total_contacts"],
        },
    )
    db.commit()


def recompute_contacts_view(db: Session) -> None:
    db.execute(text("TRUNCATE TABLE contacts_view"))
    db.execute(
        text(
            """
            WITH dedup_records AS (
              SELECT DISTINCT ON (r.phone_normalized)
                r.phone_normalized,
                r.phone,
                r.person_name,
                r.created_at
              FROM data_records r
              WHERE COALESCE(r.phone_normalized, '') <> ''
              ORDER BY r.phone_normalized, r.created_at DESC, r.id DESC
            ),
            owner_stats AS (
              SELECT
                x.phone_normalized,
                x.owner_count,
                y.owner_name AS owner_primary_name
              FROM (
                SELECT
                  r.phone_normalized,
                  COUNT(DISTINCT o.id)::int AS owner_count
                FROM data_records r
                JOIN data_owners o ON o.id = r.owner_id
                WHERE COALESCE(r.phone_normalized, '') <> ''
                GROUP BY r.phone_normalized
              ) x
              LEFT JOIN (
                SELECT DISTINCT ON (r.phone_normalized)
                  r.phone_normalized,
                  o.owner_name
                FROM data_records r
                JOIN data_owners o ON o.id = r.owner_id
                WHERE COALESCE(r.phone_normalized, '') <> ''
                ORDER BY r.phone_normalized, o.uploaded_at DESC, o.id DESC
              ) y ON y.phone_normalized = x.phone_normalized
            ),
            favorite_stats AS (
              SELECT
                r.phone_normalized,
                BOOL_OR(COALESCE(r.intimacy_checked, FALSE)) AS has_favorite
              FROM data_records r
              WHERE COALESCE(r.phone_normalized, '') <> ''
              GROUP BY r.phone_normalized
            )
            INSERT INTO contacts_view(
              phone_normalized, phone, person_name, created_at, has_favorite,
              city_county, dong, address_detail, owner_primary_name, owner_count, updated_at
            )
            SELECT
              d.phone_normalized,
              COALESCE(NULLIF(TRIM(d.phone), ''), d.phone_normalized) AS phone,
              d.person_name,
              d.created_at,
              COALESCE(f.has_favorite, FALSE) AS has_favorite,
              c.city_county,
              c.dong,
              c.address_detail,
              o.owner_primary_name,
              COALESCE(o.owner_count, 0)::int AS owner_count,
              NOW()
            FROM dedup_records d
            LEFT JOIN compare_records_view c ON c.phone_normalized = d.phone_normalized
            LEFT JOIN owner_stats o ON o.phone_normalized = d.phone_normalized
            LEFT JOIN favorite_stats f ON f.phone_normalized = d.phone_normalized
            """
        )
    )
    db.commit()


def recompute_compare_records_view(db: Session) -> None:
    db.execute(text("TRUNCATE TABLE compare_records_view"))
    db.execute(
        text(
            """
            INSERT INTO compare_records_view(
              id, full_name, birth_date, phone, phone_normalized, province,
              city_county, district, dong, address_detail, created_at, updated_at
            )
            SELECT
              id, full_name, birth_date, phone, phone_normalized, province,
              city_county, district, dong, address_detail, created_at, updated_at
            FROM compare_records
            """
        )
    )
    refresh_compare_records_view_meta(db)
    db.commit()


def refresh_compare_records_view_meta(db: Session) -> None:
    row = db.execute(
        text(
            """
            SELECT
              COUNT(*)::int AS total_count,
              COUNT(*) FILTER (WHERE COALESCE(TRIM(dong), '') <> '')::int AS dong_count,
              MAX(updated_at) AS latest_updated_at
            FROM compare_records_view
            """
        )
    ).mappings().first()

    db.execute(
        text(
            """
            INSERT INTO compare_records_view_meta(id, total_count, dong_count, latest_updated_at, refreshed_at)
            VALUES (1, :total_count, :dong_count, :latest_updated_at, NOW())
            ON CONFLICT (id)
            DO UPDATE SET
              total_count = EXCLUDED.total_count,
              dong_count = EXCLUDED.dong_count,
              latest_updated_at = EXCLUDED.latest_updated_at,
              refreshed_at = NOW()
            """
        ),
        {
            "total_count": row["total_count"] if row else 0,
            "dong_count": row["dong_count"] if row else 0,
            "latest_updated_at": row["latest_updated_at"] if row else None,
        },
    )


def upsert_compare_records_view_for_phones(
    db: Session, phone_norms: set[str] | list[str] | tuple[str, ...]
) -> None:
    phones = sorted({(p or "").strip() for p in phone_norms if (p or "").strip()})
    if not phones:
        refresh_compare_records_view_meta(db)
        return

    db.execute(
        text("DELETE FROM compare_records_view WHERE phone_normalized = ANY(:phones)"),
        {"phones": phones},
    )
    db.execute(
        text(
            """
            INSERT INTO compare_records_view(
              id, full_name, birth_date, phone, phone_normalized, province,
              city_county, district, dong, address_detail, created_at, updated_at
            )
            SELECT
              id, full_name, birth_date, phone, phone_normalized, province,
              city_county, district, dong, address_detail, created_at, updated_at
            FROM compare_records
            WHERE phone_normalized = ANY(:phones)
            """
        ),
        {"phones": phones},
    )
    refresh_compare_records_view_meta(db)


def upsert_contacts_view_for_phones(db: Session, phone_norms: set[str] | list[str] | tuple[str, ...]) -> None:
    phones = sorted({(p or "").strip() for p in phone_norms if (p or "").strip()})
    if not phones:
        return

    db.execute(
        text("DELETE FROM contacts_view WHERE phone_normalized = ANY(:phones)"),
        {"phones": phones},
    )
    db.execute(
        text(
            """
            WITH dedup_records AS (
              SELECT DISTINCT ON (r.phone_normalized)
                r.phone_normalized,
                r.phone,
                r.person_name,
                r.created_at
              FROM data_records r
              WHERE COALESCE(r.phone_normalized, '') <> ''
                AND r.phone_normalized = ANY(:phones)
              ORDER BY r.phone_normalized, r.created_at DESC, r.id DESC
            ),
            owner_stats AS (
              SELECT
                x.phone_normalized,
                x.owner_count,
                y.owner_name AS owner_primary_name
              FROM (
                SELECT
                  r.phone_normalized,
                  COUNT(DISTINCT o.id)::int AS owner_count
                FROM data_records r
                JOIN data_owners o ON o.id = r.owner_id
                WHERE COALESCE(r.phone_normalized, '') <> ''
                  AND r.phone_normalized = ANY(:phones)
                GROUP BY r.phone_normalized
              ) x
              LEFT JOIN (
                SELECT DISTINCT ON (r.phone_normalized)
                  r.phone_normalized,
                  o.owner_name
                FROM data_records r
                JOIN data_owners o ON o.id = r.owner_id
                WHERE COALESCE(r.phone_normalized, '') <> ''
                  AND r.phone_normalized = ANY(:phones)
                ORDER BY r.phone_normalized, o.uploaded_at DESC, o.id DESC
              ) y ON y.phone_normalized = x.phone_normalized
            ),
            favorite_stats AS (
              SELECT
                r.phone_normalized,
                BOOL_OR(COALESCE(r.intimacy_checked, FALSE)) AS has_favorite
              FROM data_records r
              WHERE COALESCE(r.phone_normalized, '') <> ''
                AND r.phone_normalized = ANY(:phones)
              GROUP BY r.phone_normalized
            )
            INSERT INTO contacts_view(
              phone_normalized, phone, person_name, created_at, has_favorite,
              city_county, dong, address_detail, owner_primary_name, owner_count, updated_at
            )
            SELECT
              d.phone_normalized,
              COALESCE(NULLIF(TRIM(d.phone), ''), d.phone_normalized) AS phone,
              d.person_name,
              d.created_at,
              COALESCE(f.has_favorite, FALSE) AS has_favorite,
              c.city_county,
              c.dong,
              c.address_detail,
              o.owner_primary_name,
              COALESCE(o.owner_count, 0)::int AS owner_count,
              NOW()
            FROM dedup_records d
            LEFT JOIN compare_records_view c ON c.phone_normalized = d.phone_normalized
            LEFT JOIN owner_stats o ON o.phone_normalized = d.phone_normalized
            LEFT JOIN favorite_stats f ON f.phone_normalized = d.phone_normalized
            """
        ),
        {"phones": phones},
    )
    db.commit()


def refresh_stats_if_needed(db: Session) -> None:
    summary = db.execute(
        text("SELECT refreshed_at FROM stats_view WHERE id = 1")
    ).mappings().first()
    if summary is None:
        recompute_stats_view(db)
        return

    refreshed_at = summary["refreshed_at"]
    if not isinstance(refreshed_at, datetime):
        recompute_stats_view(db)
        return

    last_kst_date = refreshed_at.astimezone(KST).date()
    today_kst_date = datetime.now(KST).date()
    if last_kst_date < today_kst_date:
        recompute_stats_view(db)


def start_daily_stats_recompute_thread() -> None:
    global _stats_thread_started
    with _stats_thread_lock:
        if _stats_thread_started:
            return
        _stats_thread_started = True

    def _worker():
        last_run_date = datetime.now(KST).date()
        while True:
            try:
                now_kst = datetime.now(KST)
                if now_kst.date() > last_run_date:
                    db = SessionLocal()
                    try:
                        recompute_contacts_view(db)
                        recompute_compare_records_view(db)
                        recompute_stats_view(db)
                    finally:
                        db.close()
                    last_run_date = now_kst.date()
            except Exception:
                # Keep the scheduler alive even if one cycle fails.
                pass
            time.sleep(60)

    thread = threading.Thread(target=_worker, name="daily-stats-recompute", daemon=True)
    thread.start()


@app.on_event("startup")
def on_startup() -> None:
    db = SessionLocal()
    try:
        ensure_tables(db)
        recompute_compare_records_view(db)
        recompute_contacts_view(db)
        recompute_stats_view(db)
        start_daily_stats_recompute_thread()
    finally:
        db.close()


@app.post("/auth/login", response_model=LoginResponse)
def login(payload: LoginRequest):
    if payload.username != ADMIN_USERNAME or payload.password != ADMIN_PASSWORD:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Invalid credentials",
        )
    return LoginResponse(access_token=ACCESS_TOKEN)


@app.get("/health")
def health_check(db: Session = Depends(get_db)):
    db.execute(text("SELECT 1"))
    return {"status": "ok"}


@app.get("/groups", response_model=list[GroupRead])
def list_groups(_: None = Depends(require_auth), db: Session = Depends(get_db)):
    rows = db.execute(
        text("SELECT id, name, created_at FROM data_groups ORDER BY id DESC")
    ).mappings()
    return [GroupRead(**row) for row in rows]


@app.post("/groups", response_model=GroupRead)
def create_group(payload: GroupCreate, _: None = Depends(require_auth), db: Session = Depends(get_db)):
    existing = db.execute(
        text("SELECT id, name, created_at FROM data_groups WHERE name = :name"),
        {"name": payload.name.strip()},
    ).mappings().first()
    if existing:
        return GroupRead(**existing)

    row = db.execute(
        text(
            """
            INSERT INTO data_groups(name)
            VALUES (:name)
            RETURNING id, name, created_at
            """
        ),
        {"name": payload.name.strip()},
    ).mappings().first()
    db.commit()
    recompute_stats_view(db)
    return GroupRead(**row)


@app.get("/groups/{group_id}/owners", response_model=list[GroupOwnerRead])
def list_group_owners(
    group_id: int,
    _: None = Depends(require_auth),
    db: Session = Depends(get_db),
):
    rows = db.execute(
        text(
            """
            SELECT
              o.id,
              o.owner_name,
              o.owner_phone,
              o.source_type,
              o.file_name,
              o.uploaded_at,
              COUNT(r.id)::int AS record_count
            FROM data_owners o
            LEFT JOIN data_records r ON r.owner_id = o.id
            WHERE o.group_id = :group_id
            GROUP BY o.id
            ORDER BY o.id DESC
            """
        ),
        {"group_id": group_id},
    ).mappings()
    return [GroupOwnerRead(**row) for row in rows]


@app.get("/stats/summary", response_model=StatsSummaryRead)
def get_stats_summary(_: None = Depends(require_auth), db: Session = Depends(get_db)):
    refresh_stats_if_needed(db)
    row = db.execute(
        text(
            f"""
            WITH unified_people AS (
              SELECT
                u.phone_normalized,
                u.phone,
                u.person_name,
                u.supporter_name,
                u.created_at,
                u.source_priority,
                u.id
              FROM (
                SELECT
                  r.phone_normalized,
                  r.phone,
                  r.person_name,
                  NULL::text AS supporter_name,
                  r.created_at,
                  r.id,
                  1 AS source_priority
                FROM data_records r
                WHERE COALESCE(r.phone_normalized, '') <> ''
                UNION ALL
                SELECT
                  s.phone_normalized,
                  s.phone,
                  NULL::text AS person_name,
                  s.supporter_name,
                  s.created_at,
                  s.id,
                  2 AS source_priority
                FROM supporter_records s
                WHERE COALESCE(s.phone_normalized, '') <> ''
              ) u
              ORDER BY u.phone_normalized, u.created_at DESC, u.source_priority ASC, u.id DESC
            ),
            unified_dedup AS (
              SELECT DISTINCT ON (phone_normalized)
                phone_normalized,
                phone,
                person_name,
                supporter_name,
                created_at
              FROM unified_people
              ORDER BY phone_normalized, created_at DESC, source_priority ASC, id DESC
            ),
            unified_meta AS (
              SELECT
                COUNT(*)::int AS unified_total_people,
                COUNT(*) FILTER (
                  WHERE {SUPPORTER_ADDRESS_MATCH_CONDITION}
                )::int AS unified_matched_with_address
              FROM unified_dedup u
              LEFT JOIN compare_records_view c ON c.phone_normalized = u.phone_normalized
            )
            SELECT
              s.total_managers,
              s.today_representatives,
              s.today_added_managers,
              s.favorite_contacts,
              s.total_contacts,
              unified_meta.unified_total_people,
              unified_meta.unified_matched_with_address,
              s.refreshed_at
            FROM stats_view s, unified_meta
            WHERE s.id = 1
            """
        )
    ).mappings().first()

    if row is None:
        recompute_stats_view(db)
        row = db.execute(
            text(
                f"""
                WITH unified_people AS (
                  SELECT
                    u.phone_normalized,
                    u.phone,
                    u.person_name,
                    u.supporter_name,
                    u.created_at,
                    u.source_priority,
                    u.id
                  FROM (
                    SELECT
                      r.phone_normalized,
                      r.phone,
                      r.person_name,
                      NULL::text AS supporter_name,
                      r.created_at,
                      r.id,
                      1 AS source_priority
                    FROM data_records r
                    WHERE COALESCE(r.phone_normalized, '') <> ''
                    UNION ALL
                    SELECT
                      s.phone_normalized,
                      s.phone,
                      NULL::text AS person_name,
                      s.supporter_name,
                      s.created_at,
                      s.id,
                      2 AS source_priority
                    FROM supporter_records s
                    WHERE COALESCE(s.phone_normalized, '') <> ''
                  ) u
                  ORDER BY u.phone_normalized, u.created_at DESC, u.source_priority ASC, u.id DESC
                ),
                unified_dedup AS (
                  SELECT DISTINCT ON (phone_normalized)
                    phone_normalized,
                    phone,
                    person_name,
                    supporter_name,
                    created_at
                  FROM unified_people
                  ORDER BY phone_normalized, created_at DESC, source_priority ASC, id DESC
                ),
                unified_meta AS (
                  SELECT
                    COUNT(*)::int AS unified_total_people,
                    COUNT(*) FILTER (
                      WHERE {SUPPORTER_ADDRESS_MATCH_CONDITION}
                    )::int AS unified_matched_with_address
                  FROM unified_dedup u
                  LEFT JOIN compare_records_view c ON c.phone_normalized = u.phone_normalized
                )
                SELECT
                  s.total_managers,
                  s.today_representatives,
                  s.today_added_managers,
                  s.favorite_contacts,
                  s.total_contacts,
                  unified_meta.unified_total_people,
                  unified_meta.unified_matched_with_address,
                  s.refreshed_at
                FROM stats_view s, unified_meta
                WHERE s.id = 1
                """
            )
        ).mappings().first()

    return StatsSummaryRead(**row)


@app.get("/stats/combined-contacts", response_model=UnifiedContactListRead)
def list_combined_contacts(
    page: int = Query(default=1, ge=1),
    scope: str = Query(default=UNIFIED_SCOPE_TOTAL),
    _: None = Depends(require_auth),
    db: Session = Depends(get_db),
):
    if scope not in {UNIFIED_SCOPE_TOTAL, UNIFIED_SCOPE_MATCHED}:
        raise HTTPException(status_code=400, detail="scope must be one of ['total', 'matched']")

    where_sql = ""
    if scope == UNIFIED_SCOPE_MATCHED:
        where_sql = f"WHERE {SUPPORTER_ADDRESS_MATCH_CONDITION}"

    page_size = 100
    meta = db.execute(
        text(
            f"""
            WITH unified_people AS (
              SELECT
                u.phone_normalized,
                u.phone,
                u.person_name,
                u.supporter_name,
                u.source,
                u.created_at,
                u.source_priority,
                u.id
              FROM (
                SELECT
                  r.phone_normalized,
                  r.phone,
                  r.person_name,
                  NULL::text AS supporter_name,
                  'acquaintance'::text AS source,
                  r.created_at,
                  1 AS source_priority,
                  r.id
                FROM data_records r
                WHERE COALESCE(r.phone_normalized, '') <> ''
                UNION ALL
                SELECT
                  s.phone_normalized,
                  s.phone,
                  NULL::text AS person_name,
                  s.supporter_name,
                  'supporter'::text AS source,
                  s.created_at,
                  2 AS source_priority,
                  s.id
                FROM supporter_records s
                WHERE COALESCE(s.phone_normalized, '') <> ''
              ) u
            ),
            unified_dedup AS (
              SELECT DISTINCT ON (phone_normalized)
                phone_normalized,
                phone,
                person_name,
                supporter_name,
                source,
                created_at
              FROM unified_people
              ORDER BY phone_normalized, created_at DESC, source_priority ASC, id DESC
            )
            SELECT COUNT(*)::int AS total_count, NOW() AS refreshed_at
            FROM unified_dedup u
            LEFT JOIN compare_records_view c ON c.phone_normalized = u.phone_normalized
            {where_sql}
            """
        )
    ).mappings().first()

    rows = db.execute(
        text(
            f"""
            WITH unified_people AS (
              SELECT
                u.phone_normalized,
                u.phone,
                u.person_name,
                u.supporter_name,
                u.source,
                u.created_at,
                u.source_priority,
                u.id
              FROM (
                SELECT
                  r.phone_normalized,
                  r.phone,
                  r.person_name,
                  NULL::text AS supporter_name,
                  'acquaintance'::text AS source,
                  r.created_at,
                  1 AS source_priority,
                  r.id
                FROM data_records r
                WHERE COALESCE(r.phone_normalized, '') <> ''
                UNION ALL
                SELECT
                  s.phone_normalized,
                  s.phone,
                  NULL::text AS person_name,
                  s.supporter_name,
                  'supporter'::text AS source,
                  s.created_at,
                  2 AS source_priority,
                  s.id
                FROM supporter_records s
                WHERE COALESCE(s.phone_normalized, '') <> ''
              ) u
            ),
            unified_dedup AS (
              SELECT DISTINCT ON (phone_normalized)
                phone_normalized,
                phone,
                person_name,
                supporter_name,
                source,
                created_at
              FROM unified_people
              ORDER BY phone_normalized, created_at DESC, source_priority ASC, id DESC
            )
            SELECT
              u.phone,
              u.person_name,
              u.supporter_name,
              u.source,
              c.city_county,
              c.district,
              c.dong,
              c.address_detail,
              u.created_at
            FROM unified_dedup u
            LEFT JOIN compare_records_view c ON c.phone_normalized = u.phone_normalized
            {where_sql}
            ORDER BY u.created_at DESC, u.phone DESC
            LIMIT :limit OFFSET :offset
            """
        ),
        {"limit": page_size, "offset": (page - 1) * page_size},
    ).mappings().all()

    total_count = int(meta["total_count"]) if meta else 0
    total_pages = max(1, (total_count + page_size - 1) // page_size)
    if total_count == 0:
        page = 1
    if page > total_pages and total_count > 0:
        page = total_pages
        rows = db.execute(
            text(
                f"""
                WITH unified_people AS (
                  SELECT
                    u.phone_normalized,
                    u.phone,
                    u.person_name,
                    u.supporter_name,
                    u.source,
                    u.created_at,
                    u.source_priority,
                    u.id
                  FROM (
                    SELECT
                      r.phone_normalized,
                      r.phone,
                      r.person_name,
                      NULL::text AS supporter_name,
                      'acquaintance'::text AS source,
                      r.created_at,
                      1 AS source_priority,
                      r.id
                    FROM data_records r
                    WHERE COALESCE(r.phone_normalized, '') <> ''
                    UNION ALL
                    SELECT
                      s.phone_normalized,
                      s.phone,
                      NULL::text AS person_name,
                      s.supporter_name,
                      'supporter'::text AS source,
                      s.created_at,
                      2 AS source_priority,
                      s.id
                    FROM supporter_records s
                    WHERE COALESCE(s.phone_normalized, '') <> ''
                  ) u
                ),
                unified_dedup AS (
                  SELECT DISTINCT ON (phone_normalized)
                    phone_normalized,
                    phone,
                    person_name,
                    supporter_name,
                    source,
                    created_at
                  FROM unified_people
                  ORDER BY phone_normalized, created_at DESC, source_priority ASC, id DESC
                )
                SELECT
                  u.phone,
                  u.person_name,
                  u.supporter_name,
                  u.source,
                  c.city_county,
                  c.district,
                  c.dong,
                  c.address_detail,
                  u.created_at
                FROM unified_dedup u
                LEFT JOIN compare_records_view c ON c.phone_normalized = u.phone_normalized
                {where_sql}
                ORDER BY u.created_at DESC, u.phone DESC
                LIMIT :limit OFFSET :offset
                """
            ),
            {"limit": page_size, "offset": (page - 1) * page_size},
        ).mappings().all()

    return UnifiedContactListRead(
        scope=scope,
        total=total_count,
        page=page,
        page_size=page_size,
        refreshed_at=meta["refreshed_at"] if meta else datetime.now(tz=KST),
        items=[UnifiedContactListItemRead(**row) for row in rows],
    )


@app.get("/stats/combined-contacts/export")
def export_combined_contacts(
    scope: str = Query(default=UNIFIED_SCOPE_TOTAL),
    _: None = Depends(require_auth),
    db: Session = Depends(get_db),
):
    if scope not in {UNIFIED_SCOPE_TOTAL, UNIFIED_SCOPE_MATCHED}:
        raise HTTPException(status_code=400, detail="scope must be one of ['total', 'matched']")

    where_sql = ""
    if scope == UNIFIED_SCOPE_MATCHED:
        where_sql = f"WHERE {SUPPORTER_ADDRESS_MATCH_CONDITION}"

    rows = db.execute(
        text(
            f"""
            WITH unified_people AS (
              SELECT
                u.phone_normalized,
                u.phone,
                u.person_name,
                u.supporter_name,
                u.source,
                u.created_at,
                u.source_priority,
                u.id
              FROM (
                SELECT
                  r.phone_normalized,
                  r.phone,
                  r.person_name,
                  NULL::text AS supporter_name,
                  'acquaintance'::text AS source,
                  r.created_at,
                  1 AS source_priority,
                  r.id
                FROM data_records r
                WHERE COALESCE(r.phone_normalized, '') <> ''
                UNION ALL
                SELECT
                  s.phone_normalized,
                  s.phone,
                  NULL::text AS person_name,
                  s.supporter_name,
                  'supporter'::text AS source,
                  s.created_at,
                  2 AS source_priority,
                  s.id
                FROM supporter_records s
                WHERE COALESCE(s.phone_normalized, '') <> ''
              ) u
            ),
            unified_dedup AS (
              SELECT DISTINCT ON (phone_normalized)
                phone_normalized,
                phone,
                person_name,
                supporter_name,
                source,
                created_at
              FROM unified_people
              ORDER BY phone_normalized, created_at DESC, source_priority ASC, id DESC
            )
            SELECT
              u.phone,
              u.person_name,
              u.supporter_name,
              u.source,
              c.city_county,
              c.district,
              c.dong,
              c.address_detail,
              u.created_at
            FROM unified_dedup u
            LEFT JOIN compare_records_view c ON c.phone_normalized = u.phone_normalized
            {where_sql}
            ORDER BY u.created_at DESC, u.phone DESC
            """
        )
    ).mappings().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "통합연락처"
    ws.append(["연락처", "지인 이름", "서포터 이름", "출처", "시(군)", "구", "동", "주소(상세)", "등록시각"])
    for row in rows:
        ws.append(
            [
                row.get("phone") or "",
                row.get("person_name") or "",
                row.get("supporter_name") or "",
                "지인" if row.get("source") == "acquaintance" else "서포터",
                row.get("city_county") or "",
                row.get("district") or "",
                row.get("dong") or "",
                row.get("address_detail") or "",
                row.get("created_at").strftime("%Y-%m-%d %H:%M:%S") if row.get("created_at") else "",
            ]
        )

    filename = "combined_total_people.xlsx" if scope == UNIFIED_SCOPE_TOTAL else "combined_matched_people.xlsx"
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def get_daily_stats_rows(db: Session):
    rows = db.execute(
        text(
            """
            WITH reps AS (
              SELECT (created_at AT TIME ZONE 'Asia/Seoul')::date AS stat_date, COUNT(*)::int AS representatives_added
              FROM data_groups
              GROUP BY 1
            ),
            managers AS (
              SELECT (uploaded_at AT TIME ZONE 'Asia/Seoul')::date AS stat_date, COUNT(*)::int AS managers_added
              FROM data_owners
              GROUP BY 1
            ),
            contacts_first AS (
              SELECT
                phone_normalized,
                MIN((created_at AT TIME ZONE 'Asia/Seoul')::date) AS first_date
              FROM data_records
              WHERE COALESCE(phone_normalized, '') <> ''
              GROUP BY phone_normalized
            ),
            contacts_daily AS (
              SELECT first_date AS stat_date, COUNT(*)::int AS contacts_added
              FROM contacts_first
              GROUP BY 1
            ),
            favorites_first AS (
              SELECT
                phone_normalized,
                MIN((created_at AT TIME ZONE 'Asia/Seoul')::date) AS first_date
              FROM data_records
              WHERE COALESCE(phone_normalized, '') <> '' AND intimacy_checked = TRUE
              GROUP BY phone_normalized
            ),
            favorites_daily AS (
              SELECT first_date AS stat_date, COUNT(*)::int AS favorite_contacts_added
              FROM favorites_first
              GROUP BY 1
            ),
            all_dates AS (
              SELECT stat_date FROM reps
              UNION
              SELECT stat_date FROM managers
              UNION
              SELECT stat_date FROM contacts_daily
              UNION
              SELECT stat_date FROM favorites_daily
            )
            SELECT
              all_dates.stat_date::text AS stat_date,
              COALESCE(reps.representatives_added, 0)::int AS representatives_added,
              COALESCE(managers.managers_added, 0)::int AS managers_added,
              COALESCE(favorites_daily.favorite_contacts_added, 0)::int AS favorite_contacts_added,
              COALESCE(contacts_daily.contacts_added, 0)::int AS contacts_added
            FROM all_dates
            LEFT JOIN reps ON reps.stat_date = all_dates.stat_date
            LEFT JOIN managers ON managers.stat_date = all_dates.stat_date
            LEFT JOIN contacts_daily ON contacts_daily.stat_date = all_dates.stat_date
            LEFT JOIN favorites_daily ON favorites_daily.stat_date = all_dates.stat_date
            ORDER BY all_dates.stat_date DESC
            """
        )
    ).mappings().all()
    return rows


@app.get("/stats/daily", response_model=list[DailyStatsRead])
def get_daily_stats(_: None = Depends(require_auth), db: Session = Depends(get_db)):
    rows = get_daily_stats_rows(db)
    return [DailyStatsRead(**row) for row in rows]


@app.get("/stats/today-managers", response_model=list[TodayManagerRead])
def get_today_managers(_: None = Depends(require_auth), db: Session = Depends(get_db)):
    rows = db.execute(
        text(
            """
            SELECT
              o.id AS owner_id,
              g.name AS group_name,
              o.owner_name,
              COUNT(r.id)::int AS contacts_count,
              COUNT(r.id) FILTER (WHERE r.intimacy_checked = TRUE)::int AS favorite_contacts_count,
              COUNT(r.id) FILTER (WHERE r.called = TRUE)::int AS called_count,
              COUNT(r.id) FILTER (WHERE r.party_member = TRUE)::int AS party_member_count,
              o.uploaded_at
            FROM data_owners o
            JOIN data_groups g ON g.id = o.group_id
            LEFT JOIN data_records r ON r.owner_id = o.id
            WHERE (o.uploaded_at AT TIME ZONE 'Asia/Seoul')::date = (NOW() AT TIME ZONE 'Asia/Seoul')::date
            GROUP BY o.id, g.name
            ORDER BY o.uploaded_at DESC, o.id DESC
            """
        )
    ).mappings().all()
    return [TodayManagerRead(**row) for row in rows]


@app.get("/stats/tree", response_model=list[TreeGroupNodeRead])
def get_tree_stats(_: None = Depends(require_auth), db: Session = Depends(get_db)):
    rows = db.execute(
        text(
            """
            WITH owner_stats AS (
              SELECT
                o.id AS owner_id,
                o.group_id,
                o.owner_name,
                COUNT(r.id)::int AS contacts_count,
                COUNT(r.id) FILTER (WHERE r.intimacy_checked = TRUE)::int AS favorite_contacts_count,
                COUNT(r.id) FILTER (WHERE r.called = TRUE)::int AS called_count,
                COUNT(r.id) FILTER (WHERE r.party_member = TRUE)::int AS party_member_count
              FROM data_owners o
              LEFT JOIN data_records r ON r.owner_id = o.id
              GROUP BY o.id, o.group_id, o.owner_name
            ),
            owner_name_dup AS (
              SELECT
                g.id AS group_id,
                COUNT(*) FILTER (WHERE os.owner_name = g.name)::int AS same_name_count
              FROM data_groups g
              LEFT JOIN owner_stats os ON os.group_id = g.id
              GROUP BY g.id
            ),
            group_totals AS (
              SELECT
                g.id AS group_id,
                g.name AS group_name,
                COALESCE(SUM(os.contacts_count), 0)::int AS children_total_contacts_count,
                COALESCE(SUM(os.favorite_contacts_count), 0)::int AS children_total_favorite_contacts_count,
                COALESCE(SUM(os.called_count), 0)::int AS children_total_called_count,
                COALESCE(SUM(os.party_member_count), 0)::int AS children_total_party_member_count,
                (COALESCE(ond.same_name_count, 0) > 1) AS ambiguous_self_children
              FROM data_groups g
              LEFT JOIN owner_stats os ON os.group_id = g.id
              LEFT JOIN owner_name_dup ond ON ond.group_id = g.id
              GROUP BY g.id, g.name, ond.same_name_count
            )
            SELECT
              gt.group_id,
              gt.group_name,
              gt.children_total_contacts_count,
              gt.children_total_favorite_contacts_count,
              gt.children_total_called_count,
              gt.children_total_party_member_count,
              gt.ambiguous_self_children,
              os.owner_id,
              os.owner_name,
              os.contacts_count,
              os.favorite_contacts_count,
              os.called_count,
              os.party_member_count,
              CASE
                WHEN gt.ambiguous_self_children = TRUE AND os.owner_name = gt.group_name THEN TRUE
                ELSE FALSE
              END AS highlight_blue
            FROM group_totals gt
            LEFT JOIN owner_stats os ON os.group_id = gt.group_id
            ORDER BY
              gt.group_id DESC,
              CASE WHEN os.owner_name = gt.group_name THEN 0 ELSE 1 END,
              os.owner_id DESC
            """
        )
    ).mappings().all()

    groups: dict[int, dict] = {}
    for row in rows:
        group_id = row["group_id"]
        if group_id not in groups:
            groups[group_id] = {
                "group_id": group_id,
                "group_name": row["group_name"],
                "children_total_contacts_count": row["children_total_contacts_count"],
                "children_total_favorite_contacts_count": row["children_total_favorite_contacts_count"],
                "children_total_called_count": row["children_total_called_count"],
                "children_total_party_member_count": row["children_total_party_member_count"],
                "ambiguous_self_children": bool(row["ambiguous_self_children"]),
                "children": [],
            }

        if row["owner_id"] is not None:
            groups[group_id]["children"].append(
                TreeOwnerNodeRead(
                    owner_id=row["owner_id"],
                    owner_name=row["owner_name"],
                    contacts_count=row["contacts_count"],
                    favorite_contacts_count=row["favorite_contacts_count"],
                    called_count=row["called_count"],
                    party_member_count=row["party_member_count"],
                    highlight_blue=bool(row["highlight_blue"]),
                )
            )

    return [TreeGroupNodeRead(**group) for group in groups.values()]


@app.get("/stats/daily/export")
def export_daily_stats(_: None = Depends(require_auth), db: Session = Depends(get_db)):
    rows = get_daily_stats_rows(db)

    wb = Workbook()
    ws = wb.active
    ws.title = "날짜별 통계"
    ws.append(
        [
            "날짜",
            "오늘 추가된 대표인원",
            "오늘 추가된 관리인원",
            "오늘 추가된 찜한 연락처 수",
            "오늘 추가된 지인 연락처 수",
        ]
    )
    for row in rows:
        ws.append(
            [
                row["stat_date"],
                row["representatives_added"],
                row["managers_added"],
                row["favorite_contacts_added"],
                row["contacts_added"],
            ]
        )

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="daily_stats.xlsx"'},
    )


@app.get("/stats/daily/details")
def get_daily_stats_details(
    stat_date: str = Query(...),
    metric: str = Query(...),
    _: None = Depends(require_auth),
    db: Session = Depends(get_db),
):
    valid_metrics = {"representatives", "managers", "favorites", "contacts"}
    if metric not in valid_metrics:
        raise HTTPException(status_code=400, detail="metric must be one of representatives, managers, favorites, contacts")

    try:
        # Validate date format YYYY-MM-DD
        datetime.strptime(stat_date, "%Y-%m-%d")
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="stat_date must be YYYY-MM-DD") from exc

    if metric == "representatives":
        rows = db.execute(
            text(
                """
                SELECT id, name, created_at
                FROM data_groups
                WHERE (created_at AT TIME ZONE 'Asia/Seoul')::date = CAST(:stat_date AS DATE)
                ORDER BY created_at DESC, id DESC
                """
            ),
            {"stat_date": stat_date},
        ).mappings().all()
        return {"metric": metric, "stat_date": stat_date, "count": len(rows), "rows": [dict(row) for row in rows]}

    if metric == "managers":
        rows = db.execute(
            text(
                """
                SELECT
                  o.id AS owner_id,
                  g.name AS group_name,
                  o.owner_name,
                  COUNT(r.id)::int AS contacts_count,
                  COUNT(r.id) FILTER (WHERE r.intimacy_checked = TRUE)::int AS favorite_contacts_count,
                  COUNT(r.id) FILTER (WHERE r.called = TRUE)::int AS called_count,
                  COUNT(r.id) FILTER (WHERE r.party_member = TRUE)::int AS party_member_count,
                  o.uploaded_at
                FROM data_owners o
                JOIN data_groups g ON g.id = o.group_id
                LEFT JOIN data_records r ON r.owner_id = o.id
                WHERE (o.uploaded_at AT TIME ZONE 'Asia/Seoul')::date = CAST(:stat_date AS DATE)
                GROUP BY o.id, g.name
                ORDER BY o.uploaded_at DESC, o.id DESC
                """
            ),
            {"stat_date": stat_date},
        ).mappings().all()
        return {"metric": metric, "stat_date": stat_date, "count": len(rows), "rows": [dict(row) for row in rows]}

    if metric == "contacts":
        rows = db.execute(
            text(
                """
                WITH contacts_first AS (
                  SELECT
                    phone_normalized,
                    MIN((created_at AT TIME ZONE 'Asia/Seoul')::date) AS first_date
                  FROM data_records
                  WHERE COALESCE(phone_normalized, '') <> ''
                  GROUP BY phone_normalized
                ),
                target AS (
                  SELECT phone_normalized
                  FROM contacts_first
                  WHERE first_date = CAST(:stat_date AS DATE)
                ),
                latest AS (
                  SELECT DISTINCT ON (r.phone_normalized)
                    r.phone_normalized,
                    r.phone,
                    r.person_name,
                    r.created_at
                  FROM data_records r
                  JOIN target t ON t.phone_normalized = r.phone_normalized
                  ORDER BY r.phone_normalized, r.created_at DESC, r.id DESC
                )
                SELECT
                  l.phone_normalized,
                  l.phone,
                  l.person_name,
                  c.city_county,
                  c.dong,
                  c.address_detail
                FROM latest l
                LEFT JOIN compare_records c ON c.phone_normalized = l.phone_normalized
                ORDER BY l.phone ASC
                """
            ),
            {"stat_date": stat_date},
        ).mappings().all()
        return {"metric": metric, "stat_date": stat_date, "count": len(rows), "rows": [dict(row) for row in rows]}

    # metric == "favorites"
    rows = db.execute(
        text(
            """
            WITH favorites_first AS (
              SELECT
                phone_normalized,
                MIN((created_at AT TIME ZONE 'Asia/Seoul')::date) AS first_date
              FROM data_records
              WHERE COALESCE(phone_normalized, '') <> '' AND intimacy_checked = TRUE
              GROUP BY phone_normalized
            ),
            target AS (
              SELECT phone_normalized
              FROM favorites_first
              WHERE first_date = CAST(:stat_date AS DATE)
            ),
            latest AS (
              SELECT DISTINCT ON (r.phone_normalized)
                r.phone_normalized,
                r.phone,
                r.person_name,
                r.created_at
              FROM data_records r
              JOIN target t ON t.phone_normalized = r.phone_normalized
              ORDER BY r.phone_normalized, r.created_at DESC, r.id DESC
            )
            SELECT
              l.phone_normalized,
              l.phone,
              l.person_name,
              c.city_county,
              c.dong,
              c.address_detail
            FROM latest l
            LEFT JOIN compare_records c ON c.phone_normalized = l.phone_normalized
            ORDER BY l.phone ASC
            """
        ),
        {"stat_date": stat_date},
    ).mappings().all()
    return {"metric": metric, "stat_date": stat_date, "count": len(rows), "rows": [dict(row) for row in rows]}


@app.post("/data/upload/async")
async def upload_data_files_async(
    group_id: int = Form(...),
    files: list[UploadFile] = File(...),
    _: None = Depends(require_auth),
    db: Session = Depends(get_db),
):
    group = db.execute(
        text("SELECT id FROM data_groups WHERE id = :id"),
        {"id": group_id},
    ).first()
    if not group:
        raise HTTPException(status_code=404, detail="Group not found")

    if not files:
        raise HTTPException(status_code=400, detail="No files provided")

    job_id = str(uuid.uuid4())
    stored_files: list[dict] = []

    for upload in files:
        filename = upload.filename or "uploaded_file"
        suffix = Path(filename).suffix.lower()
        if suffix not in {".pdf", ".xlsx"}:
            raise HTTPException(
                status_code=400,
                detail=f"Unsupported file type: {filename}. Use .pdf or .xlsx",
            )
        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
            content = await upload.read()
            tmp.write(content)
            stored_files.append(
                {
                    "filename": filename,
                    "suffix": suffix,
                    "path": tmp.name,
                }
            )

    _set_upload_job(
        job_id,
        id=job_id,
        status="queued",
        group_id=group_id,
        total_files=len(stored_files),
        processed_files=0,
        inserted_records=0,
        current_file=None,
        files=[],
        error=None,
        created_at=datetime.now(KST).isoformat(),
    )

    worker = threading.Thread(
        target=_process_data_upload_job,
        args=(job_id, group_id, stored_files),
        daemon=True,
        name=f"data-upload-{job_id[:8]}",
    )
    worker.start()

    return {
        "job_id": job_id,
        "status": "queued",
        "total_files": len(stored_files),
    }


@app.get("/data/upload/jobs/{job_id}")
def get_data_upload_job(
    job_id: str,
    _: None = Depends(require_auth),
):
    job = _get_upload_job(job_id)
    if job is None:
        raise HTTPException(status_code=404, detail="Upload job not found")
    return job


@app.get("/supporters/template")
def download_supporter_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "서포터"
    ws.append(["이름", "전화번호"])
    ws.append(["홍길동", "01012345678"])

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="supporter_template.xlsx"'},
    )


@app.post(
    "/supporters/upload",
    response_model=SupporterUploadSyncSummary | SupporterUploadAsyncQueued,
)
async def upload_supporters(
    file: UploadFile = File(...),
    _: None = Depends(require_auth),
    db: Session = Depends(get_db),
):
    filename = file.filename or "supporters.xlsx"
    suffix = Path(filename).suffix.lower()
    if suffix != ".xlsx":
        raise HTTPException(status_code=400, detail="Only .xlsx file is supported")

    with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
        content = await file.read()
        tmp.write(content)
        tmp_path = tmp.name

    try:
        rows = parse_supporter_excel_file(tmp_path)
    finally:
        try:
            os.unlink(tmp_path)
        except FileNotFoundError:
            pass
    total_rows = len(rows)

    if total_rows > 50_000:
        job_id = str(uuid.uuid4())
        _set_supporter_upload_job(
            job_id,
            id=job_id,
            status="queued",
            total_rows=total_rows,
            processed_rows=0,
            inserted=0,
            skipped_duplicate=0,
            invalid_count=0,
            error=None,
            created_at=datetime.now(KST).isoformat(),
        )
        worker = threading.Thread(
            target=_process_supporter_upload_job,
            args=(job_id, rows),
            daemon=True,
            name=f"supporter-upload-{job_id[:8]}",
        )
        worker.start()
        return SupporterUploadAsyncQueued(
            mode="async",
            job_id=job_id,
            status="queued",
            total_rows=total_rows,
        )

    result = _process_supporter_rows(db, rows)

    return SupporterUploadSyncSummary(mode="sync", **result)


@app.get("/supporters/upload/jobs/{job_id}")
def get_supporter_upload_job(
    job_id: str,
    _: None = Depends(require_auth),
):
    job = _get_supporter_upload_job(job_id)
    if job is None:
        raise HTTPException(status_code=404, detail="Upload job not found")
    return job


@app.get("/supporters/stats/summary", response_model=SupporterStatsSummaryRead)
def get_supporter_stats_summary(
    _: None = Depends(require_auth),
    db: Session = Depends(get_db),
):
    row = db.execute(
        text(
            f"""
            WITH totals AS (
              SELECT
                COUNT(*)::int AS total_supporters,
                COUNT(*) FILTER (
                  WHERE (created_at AT TIME ZONE 'Asia/Seoul')::date = (NOW() AT TIME ZONE 'Asia/Seoul')::date
                )::int AS today_added_supporters
              FROM supporter_records
            ),
            matched AS (
              SELECT
                COUNT(*)::int AS compare_matched_with_address
              FROM supporter_records s
              JOIN compare_records_view c ON c.phone_normalized = s.phone_normalized
              WHERE {SUPPORTER_ADDRESS_MATCH_CONDITION}
            )
            SELECT
              totals.total_supporters,
              totals.today_added_supporters,
              matched.compare_matched_with_address,
              NOW() AS refreshed_at
            FROM totals, matched
            """
        )
    ).mappings().first()

    return SupporterStatsSummaryRead(**row)


@app.get("/supporters/list", response_model=SupporterListRead)
def list_supporters(
    page: int = Query(default=1, ge=1),
    scope: str = Query(default=SUPPORTER_TOTAL_SCOPE),
    _: None = Depends(require_auth),
    db: Session = Depends(get_db),
):
    if scope not in {SUPPORTER_TOTAL_SCOPE, SUPPORTER_MATCHED_SCOPE}:
        raise HTTPException(status_code=400, detail="scope must be one of ['total', 'matched']")

    page_size = 100
    where_sql = ""
    if scope == SUPPORTER_MATCHED_SCOPE:
        where_sql = f"WHERE {SUPPORTER_ADDRESS_MATCH_CONDITION}"

    meta = db.execute(
        text(
            f"""
            SELECT
              COUNT(*)::int AS total_count,
              NOW() AS refreshed_at
            FROM supporter_records s
            LEFT JOIN compare_records_view c ON c.phone_normalized = s.phone_normalized
            {where_sql}
            """
        )
    ).mappings().first()

    rows = db.execute(
        text(
            f"""
            SELECT
              s.id,
              s.supporter_name,
              s.phone,
              c.full_name AS compare_full_name,
              c.city_county,
              c.district,
              c.dong,
              c.address_detail,
              s.created_at
            FROM supporter_records s
            LEFT JOIN compare_records_view c ON c.phone_normalized = s.phone_normalized
            {where_sql}
            ORDER BY s.created_at DESC, s.id DESC
            LIMIT :limit OFFSET :offset
            """
        ),
        {"limit": page_size, "offset": (page - 1) * page_size},
    ).mappings().all()

    total_count = int(meta["total_count"]) if meta else 0
    total_pages = max(1, (total_count + page_size - 1) // page_size)
    if total_count == 0:
        page = 1
    if page > total_pages and total_count > 0:
        page = total_pages
        rows = db.execute(
            text(
                f"""
                SELECT
                  s.id,
                  s.supporter_name,
                  s.phone,
                  c.full_name AS compare_full_name,
                  c.city_county,
                  c.district,
                  c.dong,
                  c.address_detail,
                  s.created_at
                FROM supporter_records s
                LEFT JOIN compare_records_view c ON c.phone_normalized = s.phone_normalized
                {where_sql}
                ORDER BY s.created_at DESC, s.id DESC
                LIMIT :limit OFFSET :offset
                """
            ),
            {"limit": page_size, "offset": (page - 1) * page_size},
        ).mappings().all()

    return SupporterListRead(
        scope=scope,
        total=total_count,
        page=page,
        page_size=page_size,
        refreshed_at=meta["refreshed_at"] if meta else datetime.now(tz=KST),
        items=[SupporterListItemRead(**row) for row in rows],
    )


@app.get("/supporters/export")
def export_supporters(
    scope: str = Query(default=SUPPORTER_TOTAL_SCOPE),
    _: None = Depends(require_auth),
    db: Session = Depends(get_db),
):
    if scope not in {SUPPORTER_TOTAL_SCOPE, SUPPORTER_MATCHED_SCOPE}:
        raise HTTPException(status_code=400, detail="scope must be one of ['total', 'matched']")

    where_sql = ""
    if scope == SUPPORTER_MATCHED_SCOPE:
        where_sql = f"WHERE {SUPPORTER_ADDRESS_MATCH_CONDITION}"

    rows = db.execute(
        text(
            f"""
            SELECT
              s.id,
              s.supporter_name,
              s.phone,
              c.full_name AS compare_full_name,
              c.city_county,
              c.district,
              c.dong,
              c.address_detail,
              s.created_at
            FROM supporter_records s
            LEFT JOIN compare_records_view c ON c.phone_normalized = s.phone_normalized
            {where_sql}
            ORDER BY s.created_at DESC, s.id DESC
            """
        )
    ).mappings().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "서포터"
    ws.append(["ID", "서포터명", "연락처", "비교군 성명", "시(군)", "구", "동", "주소(상세)", "등록시각"])

    for row in rows:
        ws.append(
            [
                row.get("id"),
                row.get("supporter_name") or "",
                row.get("phone") or "",
                row.get("compare_full_name") or "",
                row.get("city_county") or "",
                row.get("district") or "",
                row.get("dong") or "",
                row.get("address_detail") or "",
                row.get("created_at").strftime("%Y-%m-%d %H:%M:%S") if row.get("created_at") else "",
            ]
        )

    filename = "supporters_total.xlsx" if scope == SUPPORTER_TOTAL_SCOPE else "supporters_matched.xlsx"
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


JEONJU_CATEGORIES = {"all", "gap", "eul", "byeong"}


@app.post("/jeonju/upload", response_model=SupporterUploadSyncSummary)
async def upload_jeonju(
    category: str = Query(...),
    file: UploadFile = File(...),
    _: None = Depends(require_auth),
    db: Session = Depends(get_db),
):
    if category not in JEONJU_CATEGORIES:
        raise HTTPException(status_code=400, detail=f"category must be one of {sorted(JEONJU_CATEGORIES)}")
    filename = file.filename or "jeonju.xlsx"
    if Path(filename).suffix.lower() != ".xlsx":
        raise HTTPException(status_code=400, detail="Only .xlsx file is supported")

    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        content = await file.read()
        tmp.write(content)
        tmp_path = tmp.name

    try:
        rows = parse_supporter_excel_file(tmp_path)
    finally:
        try:
            os.unlink(tmp_path)
        except FileNotFoundError:
            pass

    rows_read = 0
    inserted = 0
    skipped_duplicate = 0
    invalid_count = 0
    incoming_by_phone: dict[str, dict] = {}

    for row in rows:
        rows_read += 1
        normalized = _normalize_supporter_phone(row.get("phone"))
        if not normalized:
            invalid_count += 1
            continue
        if normalized in incoming_by_phone:
            skipped_duplicate += 1
            continue
        incoming_by_phone[normalized] = {
            "category": category,
            "jeonju_name": (row.get("name") or "").strip() or None,
            "phone": format_phone(normalized),
            "phone_normalized": normalized,
        }

    if incoming_by_phone:
        existing_rows = db.execute(
            text(
                """
                SELECT phone_normalized
                FROM jeonju_records
                WHERE category = :category AND phone_normalized = ANY(:phones)
                """
            ),
            {"category": category, "phones": list(incoming_by_phone.keys())},
        ).mappings().all()
        existing_phones = {(r.get("phone_normalized") or "").strip() for r in existing_rows}
        skipped_duplicate += len(existing_phones)

        for phone_normalized, payload in incoming_by_phone.items():
            if phone_normalized in existing_phones:
                continue
            db.execute(
                text(
                    """
                    INSERT INTO jeonju_records(category, jeonju_name, phone, phone_normalized)
                    VALUES (:category, :jeonju_name, :phone, :phone_normalized)
                    ON CONFLICT (category, phone_normalized) DO NOTHING
                    """
                ),
                payload,
            )
            inserted += 1

    db.commit()
    return SupporterUploadSyncSummary(
        mode="sync",
        rows_read=rows_read,
        inserted=inserted,
        skipped_duplicate=skipped_duplicate,
        invalid_count=invalid_count,
    )


@app.get("/jeonju/contacts")
def list_jeonju_contacts(
    category: str = Query(default="all"),
    page: int = Query(default=1, ge=1),
    favorite_only: bool = Query(default=False),
    _: None = Depends(require_auth),
    db: Session = Depends(get_db),
):
    if category not in JEONJU_CATEGORIES:
        raise HTTPException(status_code=400, detail=f"category must be one of {sorted(JEONJU_CATEGORIES)}")
    page_size = 100
    offset = (page - 1) * page_size

    base_where_sql = """
        FROM jeonju_records j
        JOIN contacts_view cv ON cv.phone_normalized = j.phone_normalized
        WHERE j.category = :category
          AND COALESCE(TRIM(cv.city_county), '') <> '전주시'
          {favorite_filter}
    """
    favorite_filter = "AND COALESCE(cv.has_favorite, FALSE) = TRUE" if favorite_only else ""
    base_where_sql = base_where_sql.format(favorite_filter=favorite_filter)

    total_row = db.execute(
        text(f"SELECT COUNT(*)::int AS cnt {base_where_sql}"),
        {"category": category},
    ).mappings().first()
    total = total_row["cnt"] if total_row else 0

    list_sql = """
        SELECT
          cv.phone_normalized,
          cv.phone,
          cv.person_name AS name,
          cv.city_county,
          cv.dong,
          cv.address_detail,
          cv.owner_primary_name,
          COALESCE(cv.owner_count, 0)::int AS owner_count,
          (
            SELECT COUNT(DISTINCT r.owner_id)::int
            FROM data_records r
            WHERE r.phone_normalized = cv.phone_normalized
              AND COALESCE(r.intimacy_checked, FALSE) = TRUE
          ) AS favorite_owner_count,
          cv.created_at
        {base_where_sql}
        ORDER BY cv.created_at DESC, cv.phone ASC
        LIMIT :limit OFFSET :offset
    """.format(base_where_sql=base_where_sql)

    rows = db.execute(
        text(list_sql),
        {"category": category, "limit": page_size, "offset": offset},
    ).mappings().all()

    items = [
        {
            "phone_normalized": r["phone_normalized"],
            "phone": r["phone"],
            "name": r["name"],
            "city_county": r["city_county"],
            "dong": r["dong"],
            "address_detail": r["address_detail"],
            "owner_primary_name": r["owner_primary_name"],
            "owner_count": r["owner_count"],
            "favorite_owner_count": r["favorite_owner_count"],
            "created_at": r["created_at"].isoformat() if r["created_at"] else None,
        }
        for r in rows
    ]

    return {"total": total, "page": page, "page_size": page_size, "items": items}


@app.post("/data/upload", response_model=UploadSummary)
async def upload_data_files(
    group_id: int = Form(...),
    files: list[UploadFile] = File(...),
    _: None = Depends(require_auth),
    db: Session = Depends(get_db),
):
    group = db.execute(
        text("SELECT id FROM data_groups WHERE id = :id"),
        {"id": group_id},
    ).first()
    if not group:
        raise HTTPException(status_code=404, detail="Group not found")

    if not files:
        raise HTTPException(status_code=400, detail="No files provided")

    results: list[UploadFileResult] = []
    total_inserted = 0
    affected_phones: set[str] = set()

    for upload in files:
        filename = upload.filename or "uploaded_file"
        suffix = Path(filename).suffix.lower()

        if suffix not in {".pdf", ".xlsx"}:
            raise HTTPException(
                status_code=400,
                detail=f"Unsupported file type: {filename}. Use .pdf or .xlsx",
            )

        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
            content = await upload.read()
            tmp.write(content)
            tmp_path = tmp.name

        try:
            if suffix == ".pdf":
                parsed = parse_pdf_file(tmp_path, filename)
                source_type = "pdf"
            else:
                parsed = parse_excel_file(tmp_path, filename)
                source_type = "xlsx"
        finally:
            os.unlink(tmp_path)

        owner_phone_normalized = normalize_mobile_phone(parsed.get("owner_phone"))
        owner_phone_display = format_phone(parsed.get("owner_phone") or "") if owner_phone_normalized else None
        record_phone_normalized = [
            phone
            for phone in {
                normalize_mobile_phone(record.get("phone"))
                for record in parsed["records"]
            }
            if phone
        ]

        owner_id = _get_or_create_upload_owner(
            db,
            group_id=group_id,
            owner_name=parsed["owner_name"],
            owner_phone_normalized=owner_phone_normalized or None,
            owner_phone_display=owner_phone_display,
            source_type=source_type,
            file_name=filename,
            record_phone_normalized=record_phone_normalized,
        )

        inserted_for_file = 0
        existing_contacts = _get_existing_owner_contacts(db, owner_id)

        for record in parsed["records"]:
            normalized_phone = normalize_mobile_phone(record.get("phone"))
            if not normalized_phone:
                continue
            incoming_intimacy = record.get("intimacy_checked")
            incoming_called = record.get("called")
            incoming_party_member = record.get("party_member")

            existing = existing_contacts.get(normalized_phone)
            if existing:
                apply_intimacy = incoming_intimacy is not None and incoming_intimacy != existing["intimacy_checked"]
                apply_called = incoming_called is not None and incoming_called != existing["called"]
                apply_party_member = incoming_party_member is not None and incoming_party_member != existing["party_member"]

                if apply_intimacy or apply_called or apply_party_member:
                    db.execute(
                        text(
                            """
                            UPDATE data_records
                            SET
                              intimacy_checked = CASE WHEN :apply_intimacy THEN :intimacy_checked ELSE intimacy_checked END,
                              called = CASE WHEN :apply_called THEN :called ELSE called END,
                              party_member = CASE WHEN :apply_party_member THEN :party_member ELSE party_member END
                            WHERE id = :id
                            """
                        ),
                        {
                            "id": existing["id"],
                            "apply_intimacy": apply_intimacy,
                            "intimacy_checked": incoming_intimacy,
                            "apply_called": apply_called,
                            "called": incoming_called,
                            "apply_party_member": apply_party_member,
                            "party_member": incoming_party_member,
                        },
                    )
                    if apply_intimacy:
                        existing["intimacy_checked"] = incoming_intimacy
                    if apply_called:
                        existing["called"] = incoming_called
                    if apply_party_member:
                        existing["party_member"] = incoming_party_member
                    affected_phones.add(normalized_phone)
                continue

            inserted_row = db.execute(
                text(
                    """
                    INSERT INTO data_records(owner_id, person_name, phone, phone_normalized, intimacy_checked, called, party_member)
                    VALUES (:owner_id, :person_name, :phone, :phone_normalized, :intimacy_checked, :called, :party_member)
                    RETURNING id
                    """
                ),
                {
                    "owner_id": owner_id,
                    "person_name": record.get("name"),
                    "phone": format_phone(normalized_phone),
                    "phone_normalized": normalized_phone or None,
                    "intimacy_checked": incoming_intimacy,
                    "called": incoming_called,
                    "party_member": incoming_party_member,
                },
            ).first()
            existing_contacts[normalized_phone] = {
                "id": int(inserted_row[0]),
                "intimacy_checked": incoming_intimacy,
                "called": incoming_called,
                "party_member": incoming_party_member,
            }
            inserted_for_file += 1
            affected_phones.add(normalized_phone)

        total_inserted += inserted_for_file
        results.append(
            UploadFileResult(
                file_name=filename,
                owner_name=parsed["owner_name"],
                source_type=source_type,
                inserted_records=inserted_for_file,
            )
        )

    db.commit()
    upsert_contacts_view_for_phones(db, affected_phones)
    recompute_stats_view(db)

    return UploadSummary(
        group_id=group_id,
        processed_files=len(results),
        inserted_records=total_inserted,
        files=results,
    )


def should_replace_by_dong(existing_dong: str | None, incoming_dong: str | None) -> bool:
    old_has = bool((existing_dong or "").strip())
    new_has = bool((incoming_dong or "").strip())
    if new_has and not old_has:
        return True
    if old_has and not new_has:
        return False
    return True


def _normalize_supporter_phone(value: str | None) -> str:
    normalized = normalize_phone(value or "")
    if len(normalized) == 10 and normalized.startswith("10"):
        normalized = f"0{normalized}"
    if len(normalized) in (10, 11):
        return normalized
    return ""


def _set_upload_job(job_id: str, **updates):
    with _upload_jobs_lock:
        current = _upload_jobs.get(job_id, {})
        current.update(updates)
        _upload_jobs[job_id] = current


def _get_upload_job(job_id: str) -> dict | None:
    with _upload_jobs_lock:
        job = _upload_jobs.get(job_id)
        return dict(job) if job else None


def _set_supporter_upload_job(job_id: str, **updates):
    with _supporter_upload_jobs_lock:
        current = _supporter_upload_jobs.get(job_id, {})
        current.update(updates)
        _supporter_upload_jobs[job_id] = current


def _get_supporter_upload_job(job_id: str) -> dict | None:
    with _supporter_upload_jobs_lock:
        job = _supporter_upload_jobs.get(job_id)
        return dict(job) if job else None


def _process_supporter_rows(db: Session, rows: list[dict], progress_callback=None) -> dict:
    rows_read = 0
    inserted = 0
    skipped_duplicate = 0
    invalid_count = 0
    incoming_by_phone: dict[str, dict] = {}

    for row in rows:
        rows_read += 1
        normalized = _normalize_supporter_phone(row.get("phone"))
        if not normalized:
            invalid_count += 1
            if progress_callback and rows_read % 500 == 0:
                progress_callback(rows_read)
            continue
        if normalized in incoming_by_phone:
            skipped_duplicate += 1
            if progress_callback and rows_read % 500 == 0:
                progress_callback(rows_read)
            continue
        incoming_by_phone[normalized] = {
            "supporter_name": (row.get("name") or "").strip() or None,
            "phone": format_phone(normalized),
            "phone_normalized": normalized,
        }
        if progress_callback and rows_read % 500 == 0:
            progress_callback(rows_read)

    if incoming_by_phone:
        existing_rows = db.execute(
            text(
                """
                SELECT phone_normalized
                FROM supporter_records
                WHERE phone_normalized = ANY(:phones)
                """
            ),
            {"phones": list(incoming_by_phone.keys())},
        ).mappings().all()
        existing_phones = {(row.get("phone_normalized") or "").strip() for row in existing_rows}
        skipped_duplicate += len(existing_phones)

        for phone_normalized, payload in incoming_by_phone.items():
            if phone_normalized in existing_phones:
                continue
            db.execute(
                text(
                    """
                    INSERT INTO supporter_records(supporter_name, phone, phone_normalized)
                    VALUES (:supporter_name, :phone, :phone_normalized)
                    """
                ),
                payload,
            )
            inserted += 1

    if progress_callback:
        progress_callback(rows_read)

    db.commit()
    return {
        "rows_read": rows_read,
        "inserted": inserted,
        "skipped_duplicate": skipped_duplicate,
        "invalid_count": invalid_count,
    }


def _process_supporter_upload_job(job_id: str, rows: list[dict]):
    db = SessionLocal()
    try:
        _set_supporter_upload_job(
            job_id,
            status="processing",
            started_at=datetime.now(KST).isoformat(),
        )
        total_rows = len(rows)
        _set_supporter_upload_job(job_id, total_rows=total_rows, processed_rows=0)

        result = _process_supporter_rows(
            db,
            rows,
            progress_callback=lambda processed: _set_supporter_upload_job(job_id, processed_rows=processed),
        )
        _set_supporter_upload_job(
            job_id,
            status="completed",
            processed_rows=result["rows_read"],
            rows_read=result["rows_read"],
            inserted=result["inserted"],
            skipped_duplicate=result["skipped_duplicate"],
            invalid_count=result["invalid_count"],
            completed_at=datetime.now(KST).isoformat(),
        )
    except Exception as exc:
        db.rollback()
        _set_supporter_upload_job(
            job_id,
            status="failed",
            error=str(exc),
            completed_at=datetime.now(KST).isoformat(),
        )
    finally:
        db.close()


def _get_or_create_upload_owner(
    db: Session,
    *,
    group_id: int,
    owner_name: str,
    owner_phone_normalized: str | None,
    owner_phone_display: str | None,
    source_type: str,
    file_name: str,
    record_phone_normalized: list[str] | None = None,
) -> int:
    existing_owner = None
    normalized_owner_name = (owner_name or "").strip()

    # 1) Prefer strict match inside the same group when phone is present.
    if owner_phone_normalized:
        existing_owner = db.execute(
            text(
                """
                SELECT id
                FROM data_owners
                WHERE group_id = :group_id
                  AND owner_phone_normalized = :owner_phone_normalized
                  AND LOWER(TRIM(owner_name)) = LOWER(TRIM(:owner_name))
                ORDER BY id ASC
                LIMIT 1
                """
            ),
            {
                "group_id": group_id,
                "owner_phone_normalized": owner_phone_normalized,
                "owner_name": normalized_owner_name,
            },
        ).first()

    # 2) If owner phone is missing/ambiguous, pick existing owner that has the
    # largest overlap of contact phones among same-name owners in the same group.
    if existing_owner is None and normalized_owner_name and record_phone_normalized:
        overlap_owner = db.execute(
            text(
                """
                SELECT o.id, COUNT(*)::int AS overlap_count
                FROM data_owners o
                JOIN data_records r ON r.owner_id = o.id
                WHERE o.group_id = :group_id
                  AND LOWER(TRIM(o.owner_name)) = LOWER(TRIM(:owner_name))
                  AND r.phone_normalized = ANY(:phones)
                GROUP BY o.id
                ORDER BY overlap_count DESC, o.id DESC
                LIMIT 1
                """
            ),
            {
                "group_id": group_id,
                "owner_name": normalized_owner_name,
                "phones": record_phone_normalized,
            },
        ).first()
        if overlap_owner is not None and int(overlap_owner[1]) > 0:
            existing_owner = (int(overlap_owner[0]),)

    # 3) Last fallback: reuse latest same-name owner in the same group.
    if existing_owner is None and normalized_owner_name:
        existing_owner = db.execute(
            text(
                """
                SELECT id
                FROM data_owners
                WHERE group_id = :group_id
                  AND LOWER(TRIM(owner_name)) = LOWER(TRIM(:owner_name))
                ORDER BY id DESC
                LIMIT 1
                """
            ),
            {
                "group_id": group_id,
                "owner_name": normalized_owner_name,
            },
        ).first()

    if existing_owner:
        owner_id = int(existing_owner[0])
        db.execute(
            text(
                """
                UPDATE data_owners
                SET
                  group_id = :group_id,
                  owner_name = :owner_name,
                  owner_phone = :owner_phone,
                  owner_phone_normalized = :owner_phone_normalized,
                  source_type = :source_type,
                  file_name = :file_name
                WHERE id = :owner_id
                """
            ),
            {
                "group_id": group_id,
                "owner_name": normalized_owner_name,
                "owner_phone": owner_phone_display,
                "owner_phone_normalized": owner_phone_normalized,
                "source_type": source_type,
                "file_name": file_name,
                "owner_id": owner_id,
            },
        )
        return owner_id

    owner_row = db.execute(
        text(
            """
            INSERT INTO data_owners(group_id, owner_name, owner_phone, owner_phone_normalized, source_type, file_name)
            VALUES (:group_id, :owner_name, :owner_phone, :owner_phone_normalized, :source_type, :file_name)
            RETURNING id
            """
        ),
        {
            "group_id": group_id,
            "owner_name": normalized_owner_name or owner_name,
            "owner_phone": owner_phone_display,
            "owner_phone_normalized": owner_phone_normalized,
            "source_type": source_type,
            "file_name": file_name,
        },
    ).first()
    return int(owner_row[0])


def _get_existing_owner_contacts(db: Session, owner_id: int) -> dict[str, dict]:
    rows = db.execute(
        text(
            """
            SELECT DISTINCT ON (phone_normalized)
              id,
              phone_normalized,
              intimacy_checked,
              called,
              party_member
            FROM data_records
            WHERE owner_id = :owner_id
              AND COALESCE(phone_normalized, '') <> ''
            ORDER BY phone_normalized, id DESC
            """
        ),
        {"owner_id": owner_id},
    ).mappings().all()

    contacts: dict[str, dict] = {}
    for row in rows:
        normalized = (row.get("phone_normalized") or "").strip()
        if not normalized:
            continue
        contacts[normalized] = {
            "id": row["id"],
            "intimacy_checked": row.get("intimacy_checked"),
            "called": row.get("called"),
            "party_member": row.get("party_member"),
        }
    return contacts


def _process_data_upload_job(job_id: str, group_id: int, stored_files: list[dict]):
    db = SessionLocal()
    try:
        total_files = len(stored_files)
        inserted_records_total = 0
        file_results: list[dict] = []
        affected_phones: set[str] = set()

        _set_upload_job(
            job_id,
            status="processing",
            total_files=total_files,
            processed_files=0,
            inserted_records=0,
            current_file=None,
            files=[],
            error=None,
            started_at=datetime.now(KST).isoformat(),
        )

        for idx, file_info in enumerate(stored_files, start=1):
            filename = file_info["filename"]
            file_path = file_info["path"]
            suffix = file_info["suffix"]

            _set_upload_job(job_id, current_file=filename, processed_files=idx - 1)

            try:
                if suffix == ".pdf":
                    parsed = parse_pdf_file(file_path, filename)
                    source_type = "pdf"
                else:
                    parsed = parse_excel_file(file_path, filename)
                    source_type = "xlsx"

                owner_phone_normalized = normalize_mobile_phone(parsed.get("owner_phone"))
                owner_phone_display = (
                    format_phone(parsed.get("owner_phone") or "") if owner_phone_normalized else None
                )
                record_phone_normalized = [
                    phone
                    for phone in {
                        normalize_mobile_phone(record.get("phone"))
                        for record in parsed["records"]
                    }
                    if phone
                ]

                owner_id = _get_or_create_upload_owner(
                    db,
                    group_id=group_id,
                    owner_name=parsed["owner_name"],
                    owner_phone_normalized=owner_phone_normalized or None,
                    owner_phone_display=owner_phone_display,
                    source_type=source_type,
                    file_name=filename,
                    record_phone_normalized=record_phone_normalized,
                )

                inserted_for_file = 0
                existing_contacts = _get_existing_owner_contacts(db, owner_id)

                for record in parsed["records"]:
                    normalized_phone = normalize_mobile_phone(record.get("phone"))
                    if not normalized_phone:
                        continue
                    incoming_intimacy = record.get("intimacy_checked")
                    incoming_called = record.get("called")
                    incoming_party_member = record.get("party_member")

                    existing = existing_contacts.get(normalized_phone)
                    if existing:
                        apply_intimacy = (
                            incoming_intimacy is not None and incoming_intimacy != existing["intimacy_checked"]
                        )
                        apply_called = incoming_called is not None and incoming_called != existing["called"]
                        apply_party_member = (
                            incoming_party_member is not None and incoming_party_member != existing["party_member"]
                        )

                        if apply_intimacy or apply_called or apply_party_member:
                            db.execute(
                                text(
                                    """
                                    UPDATE data_records
                                    SET
                                      intimacy_checked = CASE WHEN :apply_intimacy THEN :intimacy_checked ELSE intimacy_checked END,
                                      called = CASE WHEN :apply_called THEN :called ELSE called END,
                                      party_member = CASE WHEN :apply_party_member THEN :party_member ELSE party_member END
                                    WHERE id = :id
                                    """
                                ),
                                {
                                    "id": existing["id"],
                                    "apply_intimacy": apply_intimacy,
                                    "intimacy_checked": incoming_intimacy,
                                    "apply_called": apply_called,
                                    "called": incoming_called,
                                    "apply_party_member": apply_party_member,
                                    "party_member": incoming_party_member,
                                },
                            )
                            if apply_intimacy:
                                existing["intimacy_checked"] = incoming_intimacy
                            if apply_called:
                                existing["called"] = incoming_called
                            if apply_party_member:
                                existing["party_member"] = incoming_party_member
                            affected_phones.add(normalized_phone)
                        continue

                    inserted_row = db.execute(
                        text(
                            """
                            INSERT INTO data_records(owner_id, person_name, phone, phone_normalized, intimacy_checked, called, party_member)
                            VALUES (:owner_id, :person_name, :phone, :phone_normalized, :intimacy_checked, :called, :party_member)
                            RETURNING id
                            """
                        ),
                        {
                            "owner_id": owner_id,
                            "person_name": record.get("name"),
                            "phone": format_phone(normalized_phone),
                            "phone_normalized": normalized_phone or None,
                            "intimacy_checked": incoming_intimacy,
                            "called": incoming_called,
                            "party_member": incoming_party_member,
                        },
                    ).first()
                    existing_contacts[normalized_phone] = {
                        "id": int(inserted_row[0]),
                        "intimacy_checked": incoming_intimacy,
                        "called": incoming_called,
                        "party_member": incoming_party_member,
                    }
                    inserted_for_file += 1
                    affected_phones.add(normalized_phone)

                inserted_records_total += inserted_for_file
                file_results.append(
                    {
                        "file_name": filename,
                        "owner_name": parsed["owner_name"],
                        "source_type": source_type,
                        "inserted_records": inserted_for_file,
                    }
                )
                _set_upload_job(
                    job_id,
                    processed_files=idx,
                    inserted_records=inserted_records_total,
                    files=list(file_results),
                )
            finally:
                try:
                    os.unlink(file_path)
                except FileNotFoundError:
                    pass

        db.commit()
        upsert_contacts_view_for_phones(db, affected_phones)
        recompute_stats_view(db)
        _set_upload_job(
            job_id,
            status="completed",
            processed_files=total_files,
            inserted_records=inserted_records_total,
            current_file=None,
            files=list(file_results),
            completed_at=datetime.now(KST).isoformat(),
        )
    except Exception as exc:
        db.rollback()
        _set_upload_job(
            job_id,
            status="failed",
            error=str(exc),
            current_file=None,
            completed_at=datetime.now(KST).isoformat(),
        )
    finally:
        db.close()


ADDRESSLESS_LABEL = "주소 없음"


def _clean_filter_value(value: str | None) -> str | None:
    if value is None:
        return None
    cleaned = value.strip()
    return cleaned if cleaned else None


def _get_contacts_payload(
    db: Session,
    page: int,
    page_size: int,
    city: str | None,
    dong: str | None,
    name_query: str | None = None,
    phone_query: str | None = None,
    favorite_only: bool = False,
):
    city_filter = _clean_filter_value(city)
    dong_filter = _clean_filter_value(dong)
    name_filter = _clean_filter_value(name_query)
    phone_filter = normalize_phone(phone_query or "")

    city_condition = ""
    if city_filter:
        city_condition = "AND city_label = :city_filter"

    dong_condition = ""
    if dong_filter:
        dong_condition = "AND dong_label = :dong_filter"

    search_condition = ""
    if name_filter:
        search_condition += " AND COALESCE(person_name, '') ILIKE :name_like"
    if phone_filter:
        search_condition += " AND COALESCE(phone_normalized, '') LIKE :phone_like"

    favorite_condition = ""
    if favorite_only:
        favorite_condition = "AND has_favorite = TRUE"

    cte = """
        WITH contacts AS (
          SELECT
            cv.phone_normalized,
            cv.phone,
            cv.person_name,
            cv.created_at,
            cv.city_county,
            cv.dong,
            cv.address_detail,
            COALESCE(cv.has_favorite, FALSE) AS has_favorite,
            cv.owner_primary_name,
            COALESCE(cv.owner_count, 0)::int AS owner_count,
            CASE
              WHEN COALESCE(TRIM(cv.city_county), '') = '' THEN :addressless
              ELSE TRIM(cv.city_county)
            END AS city_label,
            CASE
              WHEN COALESCE(TRIM(cv.dong), '') = '' THEN :addressless
              ELSE TRIM(cv.dong)
            END AS dong_label
          FROM contacts_view cv
        ),
        base_filtered AS (
          SELECT *
          FROM contacts
          WHERE 1=1
          {favorite_condition}
        ),
        city_filtered AS (
          SELECT *
          FROM base_filtered
          WHERE 1=1
          {city_condition}
        ),
        filtered_contacts AS (
          SELECT *
          FROM city_filtered
          WHERE 1=1
          {dong_condition}
          {search_condition}
        )
    """.format(
        favorite_condition=favorite_condition,
        city_condition=city_condition,
        dong_condition=dong_condition,
        search_condition=search_condition,
    )

    params = {
        "addressless": ADDRESSLESS_LABEL,
        "city_filter": city_filter,
        "dong_filter": dong_filter,
        "name_like": f"%{name_filter}%" if name_filter else None,
        "phone_like": f"%{phone_filter}%" if phone_filter else None,
        "limit": page_size,
        "offset": (page - 1) * page_size,
    }

    total_row = db.execute(
        text(
            cte
            + """
            SELECT COUNT(*)::int AS total
            FROM filtered_contacts
            """
        ),
        params,
    ).mappings().first()
    total = int(total_row["total"]) if total_row else 0

    city_rows = db.execute(
        text(
            cte
            + """
            SELECT city_label AS name, COUNT(*)::int AS count
            FROM base_filtered
            GROUP BY city_label
            ORDER BY
              CASE
                WHEN city_label = '전주시' THEN 0
                WHEN city_label = '완주군' THEN 1
                ELSE 2
              END,
              city_label ASC
            """
        ),
        params,
    ).mappings().all()

    dong_rows = db.execute(
        text(
            cte
            + """
            SELECT dong_label AS name, COUNT(*)::int AS count
            FROM city_filtered
            GROUP BY dong_label
            ORDER BY dong_label ASC
            """
        ),
        params,
    ).mappings().all()

    item_rows = db.execute(
        text(
            cte
            + """
            SELECT
              phone_normalized,
              phone,
              person_name AS name,
              CASE WHEN city_label = :addressless THEN NULL ELSE city_label END AS city_county,
              CASE WHEN dong_label = :addressless THEN NULL ELSE dong_label END AS dong,
              NULLIF(TRIM(address_detail), '') AS address_detail,
              owner_primary_name,
              owner_count,
              created_at
            FROM filtered_contacts
            ORDER BY created_at DESC, phone ASC
            LIMIT :limit OFFSET :offset
            """
        ),
        params,
    ).mappings().all()

    return total, city_rows, dong_rows, item_rows, city_filter, dong_filter


def _get_election_mapping(db: Session) -> dict[str, list[str]]:
    rows = db.execute(
        text(
            """
            SELECT district_name, dong
            FROM election_district_dongs
            ORDER BY district_name ASC, dong ASC
            """
        )
    ).mappings().all()
    mapping: dict[str, list[str]] = {}
    for row in rows:
        mapping.setdefault(row["district_name"], []).append(row["dong"])
    return mapping


@app.get("/election-districts")
def get_election_districts(
    _: None = Depends(require_auth),
    db: Session = Depends(get_db),
):
    return _get_election_mapping(db)


@app.get("/contacts", response_model=ContactListRead)
def list_contacts(
    page: int = Query(default=1, ge=1),
    city: str | None = Query(default=None),
    dong: str | None = Query(default=None),
    name: str | None = Query(default=None),
    phone: str | None = Query(default=None),
    favorite_only: bool = Query(default=False),
    _: None = Depends(require_auth),
    db: Session = Depends(get_db),
):
    page_size = 100
    total, city_rows, dong_rows, item_rows, city_filter, dong_filter = _get_contacts_payload(
        db=db,
        page=page,
        page_size=page_size,
        city=city,
        dong=dong,
        name_query=name,
        phone_query=phone,
        favorite_only=favorite_only,
    )

    if city_filter and not any(row["name"] == city_filter for row in city_rows):
        city_filter = None
        dong_filter = None
        total, city_rows, dong_rows, item_rows, _, _ = _get_contacts_payload(
            db=db,
            page=page,
            page_size=page_size,
            city=city_filter,
            dong=dong_filter,
            name_query=name,
            phone_query=phone,
            favorite_only=favorite_only,
        )

    return ContactListRead(
        total=total,
        page=page,
        page_size=page_size,
        city_categories=[ContactCategoryRead(**row) for row in city_rows],
        dong_categories=[ContactCategoryRead(**row) for row in dong_rows],
        items=[ContactListItemRead(**row) for row in item_rows],
    )


@app.get("/contacts/export")
def export_contacts(
    city: str | None = Query(default=None),
    dong: str | None = Query(default=None),
    name: str | None = Query(default=None),
    phone: str | None = Query(default=None),
    favorite_only: bool = Query(default=False),
    _: None = Depends(require_auth),
    db: Session = Depends(get_db),
):
    total, _, _, rows, _, _ = _get_contacts_payload(
        db=db,
        page=1,
        page_size=1_000_000,
        city=city,
        dong=dong,
        name_query=name,
        phone_query=phone,
        favorite_only=favorite_only,
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "지인 연락처"
    ws.append(["연락처", "이름", "시(군)", "동", "주소(상세)", "입력시각"])

    for row in rows:
        ws.append(
            [
                row["phone"],
                row["name"] or "",
                row["city_county"] or ADDRESSLESS_LABEL,
                row["dong"] or ADDRESSLESS_LABEL,
                row["address_detail"] or "",
                row["created_at"].astimezone(KST).strftime("%Y-%m-%d %H:%M:%S"),
            ]
        )

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    filename = f"contacts_{total}.xlsx"
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.get("/contacts/election", response_model=ElectionContactsRead)
def list_election_contacts(
    page: int = Query(default=1, ge=1),
    district: str | None = Query(default=None),
    unknown_only: bool = Query(default=False),
    city: str | None = Query(default=None),
    dong: str | None = Query(default=None),
    _: None = Depends(require_auth),
    db: Session = Depends(get_db),
):
    page_size = 100
    mapping = _get_election_mapping(db)
    all_mapped_dongs = sorted({dong for dongs in mapping.values() for dong in dongs})
    city_filter = _clean_filter_value(city)
    dong_filter = _clean_filter_value(dong)

    base_filters = []
    params: dict = {"limit": page_size, "offset": (page - 1) * page_size, "all_mapped_dongs": all_mapped_dongs}

    if district:
        dongs = mapping.get(district, [])
        params["district_dongs"] = dongs
        base_filters.append(
            "AND COALESCE(NULLIF(TRIM(cv.dong), ''), '__EMPTY__') = ANY(:district_dongs)"
            if dongs
            else "AND FALSE"
        )
    elif unknown_only:
        base_filters.append(
            "AND (COALESCE(NULLIF(TRIM(cv.dong), ''), '__EMPTY__') = '__EMPTY__' "
            "OR COALESCE(NULLIF(TRIM(cv.dong), ''), '__EMPTY__') <> ALL(:all_mapped_dongs))"
        )

    city_filter_sql = ""
    if city_filter:
        city_filter_sql = "AND COALESCE(NULLIF(TRIM(cv.city_county), ''), :addressless) = :city_filter"
        params["city_filter"] = city_filter

    dong_filter_sql = ""
    if dong_filter:
        dong_filter_sql = "AND COALESCE(NULLIF(TRIM(cv.dong), ''), :addressless) = :dong_filter"
        params["dong_filter"] = dong_filter
    params["addressless"] = ADDRESSLESS_LABEL

    base_where = "\n".join(base_filters) if base_filters else ""
    cte = f"""
        WITH base AS (
          SELECT cv.*
          FROM contacts_view cv
          WHERE 1=1
          {base_where}
        ),
        city_filtered AS (
          SELECT *
          FROM base cv
          WHERE 1=1
          {city_filter_sql}
        ),
        final_filtered AS (
          SELECT *
          FROM city_filtered cv
          WHERE 1=1
          {dong_filter_sql}
        )
    """

    district_rows = db.execute(
        text(
            """
            SELECT
              d.district_name AS name,
              COUNT(cv.phone_normalized)::int AS count
            FROM (
              SELECT district_name, ARRAY_AGG(dong) AS dongs
              FROM election_district_dongs
              GROUP BY district_name
            ) d
            LEFT JOIN contacts_view cv
              ON COALESCE(NULLIF(TRIM(cv.dong), ''), '__EMPTY__') = ANY(d.dongs)
            GROUP BY d.district_name
            ORDER BY d.district_name ASC
            """
        )
    ).mappings().all()

    unknown_count_row = db.execute(
        text(
            """
            SELECT COUNT(*)::int AS count
            FROM contacts_view cv
            WHERE (COALESCE(NULLIF(TRIM(cv.dong), ''), '__EMPTY__') = '__EMPTY__'
              OR COALESCE(NULLIF(TRIM(cv.dong), ''), '__EMPTY__') <> ALL(:all_mapped_dongs))
            """
        ),
        {"all_mapped_dongs": all_mapped_dongs},
    ).mappings().first()

    total_row = db.execute(
        text(
            cte
            + """
            SELECT COUNT(*)::int AS total
            FROM final_filtered
            """
        ),
        params,
    ).mappings().first()

    city_rows = db.execute(
        text(
            cte
            + """
            SELECT COALESCE(NULLIF(TRIM(city_county), ''), :addressless) AS name, COUNT(*)::int AS count
            FROM base
            GROUP BY 1
            ORDER BY
              CASE
                WHEN COALESCE(NULLIF(TRIM(city_county), ''), :addressless) = '전주시' THEN 0
                WHEN COALESCE(NULLIF(TRIM(city_county), ''), :addressless) = '완주군' THEN 1
                ELSE 2
              END,
              name ASC
            """
        ),
        params,
    ).mappings().all()

    dong_rows = db.execute(
        text(
            cte
            + """
            SELECT COALESCE(NULLIF(TRIM(dong), ''), :addressless) AS name, COUNT(*)::int AS count
            FROM city_filtered
            GROUP BY 1
            ORDER BY name ASC
            """
        ),
        params,
    ).mappings().all()

    item_rows = db.execute(
        text(
            cte
            + """
            SELECT
              phone_normalized,
              phone,
              person_name AS name,
              NULLIF(TRIM(city_county), '') AS city_county,
              NULLIF(TRIM(dong), '') AS dong,
              NULLIF(TRIM(address_detail), '') AS address_detail,
              owner_primary_name,
              owner_count,
              created_at
            FROM final_filtered
            ORDER BY created_at DESC, phone ASC
            LIMIT :limit OFFSET :offset
            """
        ),
        params,
    ).mappings().all()

    return ElectionContactsRead(
        total=int(total_row["total"]) if total_row else 0,
        page=page,
        page_size=page_size,
        districts=[ElectionDistrictRead(name=row["name"], count=row["count"], dongs=mapping.get(row["name"], [])) for row in district_rows],
        unknown_count=int(unknown_count_row["count"]) if unknown_count_row else 0,
        city_categories=[ContactCategoryRead(**row) for row in city_rows],
        dong_categories=[ContactCategoryRead(**row) for row in dong_rows],
        items=[ContactListItemRead(**row) for row in item_rows],
    )


@app.get("/contacts/election/export")
def export_election_contacts(
    district: str | None = Query(default=None),
    unknown_only: bool = Query(default=False),
    city: str | None = Query(default=None),
    dong: str | None = Query(default=None),
    _: None = Depends(require_auth),
    db: Session = Depends(get_db),
):
    mapping = _get_election_mapping(db)
    all_mapped_dongs = sorted({dong for dongs in mapping.values() for dong in dongs})
    city_filter = _clean_filter_value(city)
    dong_filter = _clean_filter_value(dong)

    base_filters = []
    params: dict = {"all_mapped_dongs": all_mapped_dongs, "addressless": ADDRESSLESS_LABEL}

    if district:
        dongs = mapping.get(district, [])
        params["district_dongs"] = dongs
        base_filters.append(
            "AND COALESCE(NULLIF(TRIM(cv.dong), ''), '__EMPTY__') = ANY(:district_dongs)"
            if dongs
            else "AND FALSE"
        )
    elif unknown_only:
        base_filters.append(
            "AND (COALESCE(NULLIF(TRIM(cv.dong), ''), '__EMPTY__') = '__EMPTY__' "
            "OR COALESCE(NULLIF(TRIM(cv.dong), ''), '__EMPTY__') <> ALL(:all_mapped_dongs))"
        )

    city_filter_sql = ""
    if city_filter:
        city_filter_sql = "AND COALESCE(NULLIF(TRIM(cv.city_county), ''), :addressless) = :city_filter"
        params["city_filter"] = city_filter

    dong_filter_sql = ""
    if dong_filter:
        dong_filter_sql = "AND COALESCE(NULLIF(TRIM(cv.dong), ''), :addressless) = :dong_filter"
        params["dong_filter"] = dong_filter

    base_where = "\n".join(base_filters) if base_filters else ""
    rows = db.execute(
        text(
            f"""
            WITH base AS (
              SELECT cv.*
              FROM contacts_view cv
              WHERE 1=1
              {base_where}
            ),
            city_filtered AS (
              SELECT *
              FROM base cv
              WHERE 1=1
              {city_filter_sql}
            ),
            final_filtered AS (
              SELECT *
              FROM city_filtered cv
              WHERE 1=1
              {dong_filter_sql}
            )
            SELECT
              phone,
              person_name AS name,
              NULLIF(TRIM(city_county), '') AS city_county,
              NULLIF(TRIM(dong), '') AS dong,
              NULLIF(TRIM(address_detail), '') AS address_detail,
              owner_primary_name,
              owner_count,
              created_at
            FROM final_filtered
            ORDER BY created_at DESC, phone ASC
            """
        ),
        params,
    ).mappings().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "선거구 연락처"
    ws.append(["연락처", "이름", "시(군)", "동", "주소(상세)", "관리인원", "입력시각"])
    for row in rows:
        ws.append(
            [
                row["phone"] or "",
                row["name"] or "",
                row["city_county"] or ADDRESSLESS_LABEL,
                row["dong"] or ADDRESSLESS_LABEL,
                row["address_detail"] or "",
                f'{row["owner_primary_name"] or "-"} ({row["owner_count"] or 0})',
                row["created_at"].astimezone(KST).strftime("%Y-%m-%d %H:%M:%S") if row.get("created_at") else "",
            ]
        )

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    filename = "election_contacts.xlsx"
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.post("/contacts/election/districts/{district_name}/dongs", response_model=ElectionDistrictRead)
def add_election_dong(
    district_name: str,
    payload: ElectionDongAddRequest,
    _: None = Depends(require_auth),
    db: Session = Depends(get_db),
):
    district_name_clean = district_name.strip()
    if not district_name_clean:
        raise HTTPException(status_code=400, detail="district_name is required")
    dong_clean = payload.dong.strip()
    if not dong_clean:
        raise HTTPException(status_code=400, detail="dong is required")

    db.execute(
        text(
            """
            INSERT INTO election_district_dongs(district_name, dong)
            VALUES (:district_name, :dong)
            ON CONFLICT (district_name, dong) DO NOTHING
            """
        ),
        {"district_name": district_name_clean, "dong": dong_clean},
    )
    db.commit()

    mapping = _get_election_mapping(db)
    count_row = db.execute(
        text(
            """
            SELECT COUNT(*)::int AS count
            FROM contacts_view
            WHERE COALESCE(NULLIF(TRIM(dong), ''), '__EMPTY__') = ANY(:dongs)
            """
        ),
        {"dongs": mapping.get(district_name_clean, [])},
    ).mappings().first()

    return ElectionDistrictRead(
        name=district_name_clean,
        count=int(count_row["count"]) if count_row else 0,
        dongs=mapping.get(district_name_clean, []),
    )


@app.delete("/contacts/election/districts/{district_name}/dongs/{dong}", response_model=ElectionDistrictRead)
def delete_election_dong(
    district_name: str,
    dong: str,
    _: None = Depends(require_auth),
    db: Session = Depends(get_db),
):
    district_name_clean = district_name.strip()
    dong_clean = dong.strip()
    if not district_name_clean or not dong_clean:
        raise HTTPException(status_code=400, detail="district_name and dong are required")

    db.execute(
        text(
            """
            DELETE FROM election_district_dongs
            WHERE district_name = :district_name AND dong = :dong
            """
        ),
        {"district_name": district_name_clean, "dong": dong_clean},
    )
    db.commit()

    mapping = _get_election_mapping(db)
    count_row = db.execute(
        text(
            """
            SELECT COUNT(*)::int AS count
            FROM contacts_view
            WHERE COALESCE(NULLIF(TRIM(dong), ''), '__EMPTY__') = ANY(:dongs)
            """
        ),
        {"dongs": mapping.get(district_name_clean, [])},
    ).mappings().first()

    return ElectionDistrictRead(
        name=district_name_clean,
        count=int(count_row["count"]) if count_row else 0,
        dongs=mapping.get(district_name_clean, []),
    )


@app.get("/contacts/{phone_normalized}/owners", response_model=list[ContactOwnerRead])
def list_contact_owners(
    phone_normalized: str,
    favorite_only: bool = Query(default=False),
    _: None = Depends(require_auth),
    db: Session = Depends(get_db),
):
    normalized = normalize_phone(phone_normalized)
    if not normalized:
        raise HTTPException(status_code=400, detail="Invalid phone")

    favorite_filter = "AND COALESCE(r.intimacy_checked, FALSE) = TRUE" if favorite_only else ""

    rows = db.execute(
        text(
            """
            SELECT
              o.id,
              o.owner_name,
              o.owner_phone,
              g.name AS group_name,
              o.source_type,
              o.file_name,
              o.uploaded_at,
              COUNT(r.id)::int AS contact_record_count
            FROM data_records r
            JOIN data_owners o ON o.id = r.owner_id
            JOIN data_groups g ON g.id = o.group_id
            WHERE r.phone_normalized = :phone_normalized
              {favorite_filter}
            GROUP BY o.id, g.name
            ORDER BY o.uploaded_at DESC, o.id DESC
            """
            .format(favorite_filter=favorite_filter)
        ),
        {"phone_normalized": normalized},
    ).mappings().all()

    return [ContactOwnerRead(**row) for row in rows]


@app.get("/jeonju/contacts/export")
def export_jeonju_contacts(
    category: str = Query(default="all"),
    favorite_only: bool = Query(default=False),
    _: None = Depends(require_auth),
    db: Session = Depends(get_db),
):
    if category not in JEONJU_CATEGORIES:
        raise HTTPException(status_code=400, detail=f"category must be one of {sorted(JEONJU_CATEGORIES)}")

    favorite_filter = "AND COALESCE(cv.has_favorite, FALSE) = TRUE" if favorite_only else ""
    rows = db.execute(
        text(
            """
            SELECT
              cv.phone,
              cv.person_name AS name,
              NULLIF(TRIM(cv.city_county), '') AS city_county,
              NULLIF(TRIM(cv.dong), '') AS dong,
              NULLIF(TRIM(cv.address_detail), '') AS address_detail,
              cv.created_at
            FROM jeonju_records j
            JOIN contacts_view cv ON cv.phone_normalized = j.phone_normalized
            WHERE j.category = :category
              AND COALESCE(TRIM(cv.city_county), '') <> '전주시'
              {favorite_filter}
            ORDER BY cv.created_at DESC, cv.phone ASC
            """
            .format(favorite_filter=favorite_filter)
        ),
        {"category": category},
    ).mappings().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "전주시 미분류 연락처"
    ws.append(["연락처", "이름", "시(군)", "동", "주소(상세)", "입력시각"])
    for row in rows:
        ws.append(
            [
                row["phone"] or "",
                row["name"] or "",
                row["city_county"] or ADDRESSLESS_LABEL,
                row["dong"] or ADDRESSLESS_LABEL,
                row["address_detail"] or "",
                row["created_at"].astimezone(KST).strftime("%Y-%m-%d %H:%M:%S") if row.get("created_at") else "",
            ]
        )

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    filename = f"jeonju_contacts_{category}.xlsx"
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.get("/owners/{owner_id}", response_model=OwnerDetailRead)
def get_owner_detail(
    owner_id: int,
    _: None = Depends(require_auth),
    db: Session = Depends(get_db),
):
    owner, record_rows = _fetch_owner_detail_rows(db, owner_id)
    return OwnerDetailRead(
        **owner,
        records=[OwnerRecordRead(**row) for row in record_rows],
    )


@app.delete("/owners/{owner_id}")
def delete_owner(
    owner_id: int,
    _: None = Depends(require_auth),
    db: Session = Depends(get_db),
):
    owner_row = db.execute(
        text(
            """
            SELECT id
            FROM data_owners
            WHERE id = :owner_id
            """
        ),
        {"owner_id": owner_id},
    ).first()
    if owner_row is None:
        raise HTTPException(status_code=404, detail="Owner not found")

    phone_rows = db.execute(
        text(
            """
            SELECT DISTINCT phone_normalized
            FROM data_records
            WHERE owner_id = :owner_id
              AND COALESCE(phone_normalized, '') <> ''
            """
        ),
        {"owner_id": owner_id},
    ).mappings().all()
    affected_phones = {(row.get("phone_normalized") or "").strip() for row in phone_rows if row.get("phone_normalized")}

    db.execute(
        text(
            """
            DELETE FROM data_owners
            WHERE id = :owner_id
            """
        ),
        {"owner_id": owner_id},
    )
    db.commit()

    upsert_contacts_view_for_phones(db, affected_phones)
    recompute_stats_view(db)
    return {"deleted": True, "owner_id": owner_id}


def _fetch_owner_detail_rows(db: Session, owner_id: int):
    owner = db.execute(
        text(
            """
            SELECT
              o.id,
              g.name AS group_name,
              o.owner_name,
              o.owner_phone,
              o.source_type,
              o.file_name,
              o.uploaded_at
            FROM data_owners o
            JOIN data_groups g ON g.id = o.group_id
            WHERE o.id = :owner_id
            """
        ),
        {"owner_id": owner_id},
    ).mappings().first()
    if owner is None:
        raise HTTPException(status_code=404, detail="Owner not found")

    record_rows = db.execute(
        text(
            """
            SELECT
              r.id,
              r.person_name,
              r.phone,
              c.province,
              c.city_county,
              c.district,
              c.dong,
              c.address_detail,
              r.intimacy_checked,
              r.called,
              r.party_member,
              r.created_at
            FROM data_records r
            LEFT JOIN compare_records c ON c.phone_normalized = r.phone_normalized
            WHERE r.owner_id = :owner_id
            ORDER BY
              CASE WHEN r.intimacy_checked = TRUE THEN 0 ELSE 1 END,
              r.created_at DESC,
              r.id DESC
            """
        ),
        {"owner_id": owner_id},
    ).mappings().all()
    return owner, record_rows


@app.get("/owners/{owner_id}/export")
def export_owner_detail(
    owner_id: int,
    _: None = Depends(require_auth),
    db: Session = Depends(get_db),
):
    owner, record_rows = _fetch_owner_detail_rows(db, owner_id)

    wb = Workbook()
    ws = wb.active
    ws.title = "지인 목록"
    ws.append(["이름", "연락처", "도", "시(군)", "구", "동", "주소(상세)", "찜", "전화", "당원체크"])
    favorite_fill = PatternFill(fill_type="solid", fgColor="FEF3C7")
    for row in record_rows:
        values = [
            row.get("person_name") or "",
            row.get("phone") or "",
            row.get("province") or "",
            row.get("city_county") or "",
            row.get("district") or "",
            row.get("dong") or "",
            row.get("address_detail") or "",
            "O" if row.get("intimacy_checked") else "X",
            "O" if row.get("called") else "X",
            "O" if row.get("party_member") else "X",
        ]
        ws.append(values)
        if row.get("intimacy_checked"):
            excel_row = ws.max_row
            for col in range(1, len(values) + 1):
                ws.cell(row=excel_row, column=col).fill = favorite_fill

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    safe_group = (owner.get("group_name") or "group").replace(" ", "_")
    safe_owner = (owner.get("owner_name") or "owner").replace(" ", "_")
    filename = f"tree_contacts_{safe_group}_{safe_owner}_{owner_id}.xlsx"
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.get("/compare-records/template")
def download_compare_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "비교군"

    headers = ["성명", "생년월일", "연락처", "도", "시(군)", "구", "동", "주소 (상세)"]
    ws.append(headers)
    ws.append(["홍길동", "1980-01-01", "01012345678", "경기도", "수원시", "팔달구", "인계동", "123-45"])

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="compare_template.xlsx"'},
    )


@app.post("/convert/jeonju-upload")
async def convert_jeonju_upload_file(
    file: UploadFile = File(...),
    _: None = Depends(require_auth),
):
    filename = file.filename or "jeonju_upload.xlsx"
    suffix = Path(filename).suffix.lower()
    if suffix != ".xlsx":
        raise HTTPException(status_code=400, detail="Only .xlsx file is supported")

    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        content = await file.read()
        tmp.write(content)
        tmp_path = tmp.name

    try:
        rows = parse_jeonju_upload_excel_file(tmp_path)
    finally:
        os.unlink(tmp_path)

    if not rows:
        raise HTTPException(status_code=400, detail="변환 가능한 데이터가 없습니다. 시트/헤더를 확인해주세요.")

    wb = Workbook()
    ws = wb.active
    ws.title = "비교군"
    ws.append(["성명", "생년월일", "연락처", "도", "시(군)", "구", "동", "주소 (상세)"])
    for row in rows:
        ws.append(
            [
                row.get("full_name") or "",
                row.get("birth_date") or "",
                row.get("phone") or "",
                row.get("province") or "",
                row.get("city_county") or "",
                row.get("district") or "",
                row.get("dong") or "",
                row.get("address_detail") or "",
            ]
        )

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="jeonju_converted_compare.xlsx"'},
    )


@app.post("/compare-records/upload", response_model=CompareUploadSummary)
async def upload_compare_records(
    files: list[UploadFile] = File(...),
    _: None = Depends(require_auth),
    db: Session = Depends(get_db),
):
    if not files:
        raise HTTPException(status_code=400, detail="No files provided")

    incoming_by_phone: dict[str, dict] = {}
    error_batch_id = str(uuid.uuid4())
    invalid_rows: list[dict] = []
    rows_read = 0
    skipped_no_phone = 0

    for upload in files:
        filename = upload.filename or "compare.xlsx"
        suffix = Path(filename).suffix.lower()
        if suffix != ".xlsx":
            raise HTTPException(status_code=400, detail=f"Unsupported file type: {filename}. Use .xlsx")

        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
            content = await upload.read()
            tmp.write(content)
            tmp_path = tmp.name

        try:
            rows = parse_compare_excel_file(tmp_path)
        finally:
            os.unlink(tmp_path)

        for row in rows:
            rows_read += 1
            raw_phone = (row.get("phone") or "").strip()
            normalized = normalize_phone(row.get("phone"))
            if len(normalized) == 10 and normalized.startswith("10"):
                normalized = f"0{normalized}"
            if not normalized:
                skipped_no_phone += 1
                invalid_rows.append(
                    {
                        "batch_id": error_batch_id,
                        "file_name": filename,
                        "row_number": row.get("_row_number"),
                        "full_name": (row.get("full_name") or "").strip() or None,
                        "phone_raw": raw_phone or None,
                        "reason": "연락처 없음 또는 숫자 추출 불가",
                    }
                )
                continue

            if len(normalized) not in (10, 11):
                skipped_no_phone += 1
                invalid_rows.append(
                    {
                        "batch_id": error_batch_id,
                        "file_name": filename,
                        "row_number": row.get("_row_number"),
                        "full_name": (row.get("full_name") or "").strip() or None,
                        "phone_raw": raw_phone or None,
                        "reason": "연락처 자리수 오류(10~11자리 필요)",
                    }
                )
                continue

            candidate = {
                "full_name": (row.get("full_name") or "").strip() or None,
                "birth_date": (row.get("birth_date") or "").strip() or None,
                "phone": format_phone(normalized),
                "phone_normalized": normalized,
                "province": (row.get("province") or "").strip() or None,
                "city_county": (row.get("city_county") or "").strip() or None,
                "district": (row.get("district") or "").strip() or None,
                "dong": (row.get("dong") or "").strip() or None,
                "address_detail": (row.get("address_detail") or "").strip() or None,
            }

            existing = incoming_by_phone.get(normalized)
            if existing is None:
                incoming_by_phone[normalized] = candidate
                continue
            if should_replace_by_dong(existing.get("dong"), candidate.get("dong")):
                incoming_by_phone[normalized] = candidate

    inserted = 0
    updated = 0
    touched_compare_phones: set[str] = set()

    for normalized, incoming in incoming_by_phone.items():
        existing_db = db.execute(
            text(
                """
                SELECT id, dong
                FROM compare_records
                WHERE phone_normalized = :phone_normalized
                """
            ),
            {"phone_normalized": normalized},
        ).mappings().first()

        if existing_db is None:
            db.execute(
                text(
                    """
                    INSERT INTO compare_records(
                      full_name, birth_date, phone, phone_normalized, province, city_county, district, dong, address_detail
                    )
                    VALUES(
                      :full_name, :birth_date, :phone, :phone_normalized, :province, :city_county, :district, :dong, :address_detail
                    )
                    """
                ),
                incoming,
            )
            inserted += 1
            touched_compare_phones.add(normalized)
            continue

        if should_replace_by_dong(existing_db.get("dong"), incoming.get("dong")):
            db.execute(
                text(
                    """
                    UPDATE compare_records
                    SET
                      full_name = :full_name,
                      birth_date = :birth_date,
                      phone = :phone,
                      province = :province,
                      city_county = :city_county,
                      district = :district,
                      dong = :dong,
                      address_detail = :address_detail,
                      updated_at = NOW()
                    WHERE id = :id
                    """
                ),
                {**incoming, "id": existing_db["id"]},
            )
            updated += 1
            touched_compare_phones.add(normalized)

    for error_row in invalid_rows:
        db.execute(
            text(
                """
                INSERT INTO compare_upload_errors(batch_id, file_name, row_number, full_name, phone_raw, reason)
                VALUES (:batch_id, :file_name, :row_number, :full_name, :phone_raw, :reason)
                """
            ),
            error_row,
        )

    upsert_compare_records_view_for_phones(db, touched_compare_phones)
    upsert_contacts_view_for_phones(db, touched_compare_phones)
    db.commit()

    return CompareUploadSummary(
        processed_files=len(files),
        rows_read=rows_read,
        inserted=inserted,
        updated=updated,
        skipped_no_phone=skipped_no_phone,
        invalid_count=len(invalid_rows),
        error_batch_id=error_batch_id if invalid_rows else None,
    )


@app.get("/compare-records/upload-errors/{batch_id}")
def download_compare_upload_errors(
    batch_id: str,
    _: None = Depends(require_auth),
    db: Session = Depends(get_db),
):
    rows = db.execute(
        text(
            """
            SELECT file_name, row_number, full_name, phone_raw, reason, created_at
            FROM compare_upload_errors
            WHERE batch_id = :batch_id
            ORDER BY id ASC
            """
        ),
        {"batch_id": batch_id},
    ).mappings().all()

    if not rows:
        raise HTTPException(status_code=404, detail="No invalid rows found for this batch")

    output = StringIO()
    writer = DictWriter(
        output,
        fieldnames=["file_name", "row_number", "full_name", "phone_raw", "reason", "created_at"],
    )
    writer.writeheader()
    for row in rows:
        writer.writerow(dict(row))

    data = "\ufeff" + output.getvalue()
    stream = BytesIO(data.encode("utf-8"))
    return StreamingResponse(
        stream,
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="compare_upload_errors_{batch_id}.csv"'},
    )


@app.get("/compare-records/export")
def export_compare_records(
    address_contains: str = Query(default=None),
    district_name: str = Query(default=None),
    _: None = Depends(require_auth),
    db: Session = Depends(get_db),
):
    if district_name:
        mapping = _get_election_mapping(db)
        dongs = mapping.get(district_name)
        if not dongs:
            raise HTTPException(status_code=400, detail=f"선거구 '{district_name}'를 찾을 수 없습니다.")
        conditions = " OR ".join([f"COALESCE(dong, '') ILIKE :dong_{i}" for i in range(len(dongs))])
        params = {f"dong_{i}": f"%{dong}%" for i, dong in enumerate(dongs)}
        rows = db.execute(
            text(
                f"""
                SELECT
                  id, full_name, birth_date, phone, province, city_county, district, dong, address_detail, created_at, updated_at
                FROM compare_records_view
                WHERE {conditions}
                ORDER BY updated_at DESC, id DESC
                """
            ),
            params,
        ).mappings().all()
        filename = "compare_records_district.xlsx"
    else:
        keyword = (address_contains or "").strip()
        if not keyword:
            raise HTTPException(status_code=400, detail="address_contains 또는 district_name이 필요합니다.")
        rows = db.execute(
            text(
                """
                SELECT
                  id, full_name, birth_date, phone, province, city_county, district, dong, address_detail, created_at, updated_at
                FROM compare_records_view
                WHERE
                  COALESCE(city_county, '') ILIKE :keyword
                  OR COALESCE(district, '') ILIKE :keyword
                  OR COALESCE(dong, '') ILIKE :keyword
                  OR COALESCE(address_detail, '') ILIKE :keyword
                ORDER BY updated_at DESC, id DESC
                """
            ),
            {"keyword": f"%{keyword}%"},
        ).mappings().all()
        safe_keyword = "".join(ch for ch in keyword if ch.isascii() and ch.isalnum()) or "keyword"
        filename = f"compare_records_{safe_keyword}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "비교군"
    ws.append(["ID", "성명", "생년월일", "연락처", "도", "시(군)", "구", "동", "주소(상세)", "생성시각", "수정시각"])

    for row in rows:
        ws.append(
            [
                row.get("id"),
                row.get("full_name") or "",
                row.get("birth_date") or "",
                row.get("phone") or "",
                row.get("province") or "",
                row.get("city_county") or "",
                row.get("district") or "",
                row.get("dong") or "",
                row.get("address_detail") or "",
                row.get("created_at").strftime("%Y-%m-%d %H:%M:%S") if row.get("created_at") else "",
                row.get("updated_at").strftime("%Y-%m-%d %H:%M:%S") if row.get("updated_at") else "",
            ]
        )

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.get("/compare-records", response_model=CompareRecordListRead)
def list_compare_records(
    page: int = Query(default=1, ge=1),
    address_contains: str | None = Query(default=None),
    district_name: str | None = Query(default=None),
    _: None = Depends(require_auth),
    db: Session = Depends(get_db),
):
    page_size = 100
    where_sql = ""
    query_params: dict[str, object] = {}
    if district_name:
        mapping = _get_election_mapping(db)
        dongs = mapping.get(district_name)
        if not dongs:
            raise HTTPException(status_code=400, detail=f"선거구 '{district_name}'를 찾을 수 없습니다.")
        where_sql = "WHERE " + " OR ".join([f"COALESCE(dong, '') ILIKE :dong_{i}" for i in range(len(dongs))])
        query_params = {f"dong_{i}": f"%{dong}%" for i, dong in enumerate(dongs)}
    else:
        keyword = (address_contains or "").strip()
        if keyword:
            where_sql = """
                WHERE
                  COALESCE(city_county, '') ILIKE :keyword
                  OR COALESCE(district, '') ILIKE :keyword
                  OR COALESCE(dong, '') ILIKE :keyword
                  OR COALESCE(address_detail, '') ILIKE :keyword
            """
            query_params = {"keyword": f"%{keyword}%"}

    meta = db.execute(
        text(
            f"""
            SELECT
              COUNT(*)::int AS total_count,
              COUNT(*) FILTER (WHERE COALESCE(TRIM(dong), '') <> '')::int AS dong_count,
              MAX(updated_at) AS latest_updated_at,
              NOW() AS refreshed_at
            FROM compare_records_view
            {where_sql}
            """
        ),
        query_params,
    ).mappings().first()

    rows = db.execute(
        text(
            f"""
            SELECT id, full_name, birth_date, phone, province, city_county, district, dong, address_detail, created_at, updated_at
            FROM compare_records_view
            {where_sql}
            ORDER BY updated_at DESC, id DESC
            LIMIT :limit OFFSET :offset
            """
        ),
        {**query_params, "limit": page_size, "offset": (page - 1) * page_size},
    ).mappings().all()

    total_count = int(meta["total_count"]) if meta else 0
    total_pages = max(1, (total_count + page_size - 1) // page_size)
    if total_count == 0:
        page = 1
    if page > total_pages and total_count > 0:
        page = total_pages
        rows = db.execute(
            text(
                f"""
                SELECT id, full_name, birth_date, phone, province, city_county, district, dong, address_detail, created_at, updated_at
                FROM compare_records_view
                {where_sql}
                ORDER BY updated_at DESC, id DESC
                LIMIT :limit OFFSET :offset
                """
            ),
            {**query_params, "limit": page_size, "offset": (page - 1) * page_size},
        ).mappings().all()

    return CompareRecordListRead(
        total=total_count,
        dong_count=int(meta["dong_count"]) if meta else 0,
        page=page,
        page_size=page_size,
        latest_updated_at=meta["latest_updated_at"] if meta else None,
        refreshed_at=meta["refreshed_at"] if meta else None,
        items=[CompareRecordRead(**row) for row in rows],
    )
